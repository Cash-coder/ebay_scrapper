# cd "C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\ebay\ebay_test\ebay_test\spiders"
# scrapy crawl ebay_tester -o crawler_output.json
# scrapy crawl ebay_tester -o crawler_output.json > log.txt
# move crawler_output.json "C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder" -force

import traceback
import scrapy
import csv
from scrapy.http.request import Request
from scrapy.spiders import Spider
from scrapy.selector import Selector
from scrapy.loader.processors import MapCompose
from scrapy.spiders import CrawlSpider, Rule
from scrapy.linkextractors import LinkExtractor
from bs4 import BeautifulSoup as bs4
from scrapy.http import HtmlResponse
import logging
from time import sleep

# logging.basicConfig(level=logging.DEBUG)
# logging.basicConfig(level=logging.INFO)
# logging.basicConfig(level=logging.WARNING)

PRODUCTS_DB     = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\PRODS_DB.xlsx"
GAPS_FILE       = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\gaps_file.xlsx"
OUTPUT_FILENAME = 'crawler_output.json' # used only to delete old output, not to create the new one. New creates from cmd when launching spider

# IN GAPS_FILE
GAPS_TITLE_COL  = 0
GAPS_PROD_STATE = 1
GAPS_MODEL_COL  = 2
GAPS_ATTR_1_COL = 3
GAPS_ATTR_2_COL = 4
GAPS_QUANTITY   = 5
GAPS_CATEGORY   = 6
GAPS_AVAILABLE_COLORS= 7
GAPS_MIN_PRICE  = 8
GAPS_EBAY_ID    = 'J2' #they're all in the same cell

# In prods_DB
PRODSDB_EBAY_ID = 'R'

#used to avoid process products we already have
# go to prods_db, get all the ebay id's
def get_ebay_id_list():
    from openpyxl import load_workbook
    #gaps_file
    wb = load_workbook(PRODUCTS_DB)
    ws = wb.active

    ebay_ids = []

    for cell in ws[PRODSDB_EBAY_ID]:
        if cell.value != None:
            # print(cell.value)
            ebay_ids.append(cell.value)
    print(f'ebay id len: {len(ebay_ids)}')
    return ebay_ids


def get_queries():
    """ this takes a file named queries.csv and creates url's to search
    by the scrapper """
    from openpyxl import load_workbook

    #gaps_file
    wb = load_workbook(GAPS_FILE)
    ws = wb.active

    #get all the rows
    for row in ws.iter_rows(values_only=True, min_row=3):
        #query_title = iphone 12 // later I add: <iphone 12 64 gb verde> // and finally I create a different var query_search_title = query_title + prod_state// this is used to include "reacondicionado" in query, but not like mandatory to pass the filter
        query_title =   row[GAPS_TITLE_COL]
        query_prod_state=row[GAPS_PROD_STATE]
        query_model =   row[GAPS_MODEL_COL]
        query_attribute_1=row[GAPS_ATTR_1_COL] #FI: GB
        query_attribute_2=row[GAPS_ATTR_2_COL] #FI: color
        query_quantity= row[GAPS_QUANTITY]
        target_category= row[GAPS_CATEGORY]
        query_price=    row[GAPS_MIN_PRICE] #filter price, like min 300€, not prods of 5€
        available_colors=   row[GAPS_AVAILABLE_COLORS] # FI: iphone 12 != iphone 12 PRO // exclue those that has MAX or PRO in title

        if query_title == None : continue
        
        # print(
        #     'query_title: ',query_title, '\n',
        #     'query_model: ',query_model, '\n',
        #     'query_attribute_1: ',query_attribute_1, '\n',
        #     'query_attribute_2: ',query_attribute_2, '\n',
        #     'query_quantity: ',query_quantity, '\n',
        #     'target_category: ',target_category, '\n',
        #     'query_price: ',query_price, '\n',
        #     'available_colors: ',available_colors, '\n'
        #     'query_prod_state', query_prod_state
        # )

        #OLD when I targeted for colors, now I target only for GB and apply colors later
        #if query attr (color) exists, append it to title: iphone -> iphone amarillo
        # if query_attribute_1 != None and query_attribute_1 != '':
        #     query_title = query_model + ' ' + query_attribute_1
        #     #if it's a tuple like negro,grafito. Take the 1º as attr and append it query_title and the 2º as var alt_attr_1
        #     if ',' in query_attribute_1:
        #         query_attribute_1 = query_attribute_1.split(',')
        #         query_attribute_1 = query_attribute_1[0]
        #         query_alt_attr_1 = query_attribute_1[1]
        #     else: #apply deafault values in case some attr dones't exist, to append to the list cleanly
        #         query_alt_attr_1 = 'default'

        # if query_attribute_2 != None and query_attribute_2 != '':
        #     query_title = query_title + ' ' + query_attribute_2
        #     if ',' in query_attribute_2:
        #         query_attribute_2 = query_attribute_2.split(',')
        #         query_attribute_2 = query_attribute_2[0]
        #         query_alt_attr_2 = query_attribute_2[1]
        #     else:
        #         query_alt_attr_2 = 'default'

        entry = {
        'query_title':query_title,
        'query_attribute_1':query_attribute_1,
        'query_attribute_2':query_attribute_2,
        'query_quantity':query_quantity,
        'target_category':target_category,
        'query_price':query_price,
        'available_colors':available_colors,
        'query_model':query_model,
        'query_prod_state':query_prod_state
        }

        yield entry


def create_url(query_title, query_prod_state):
    # in gaps file is specified new or not new. Include some kws in the query to maximize results fo each
    if query_prod_state == 'not_new':
        query_title = query_title + ' usado'
    elif query_prod_state == 'new':
        query_title = query_title + ' nuevo'
    
    # NEW using search instead of picking categories
    first_chunk     = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
    formatted_query = query_title.replace(' ','+')
    second_chunk    = '&_sacat=0&_ipg=60&rt=nc&LH_BIN=1' #_ipg = n of articles/page
    # second_chunk    = '&_sacat=0&_ipg=120&rt=nc&LH_BIN=1' #_ipg = n of articles/page

    # second_chunk    = '&_sacat=0&_ipg=120' #_ipg = n of articles/page
    # second_chunk    = '&_sacat=0&rt=nc&LH_BIN=1&_ipg=120' #_ipg = n of articles/page
    # second_chunk    = '&_sacat=0&rt=nc&LH_BIN=1&_ipg=200' #_ipg = n of articles/page

    query_url = first_chunk + formatted_query + second_chunk
    print(f'------------- {query_url}')

    #WARNING! use url with 200 results and buy only option
    
    #OLD: url based on ebay category
    #use one or other chunks based on target category
    # if query_category == 'smartphones':
    #     first_url_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
    #     second_url_chunk = '&_sacat=9355&LH_TitleDesc=0&rt=nc&LH_BIN=1&_ipg=200'

    return query_url

def delete_old_output():
    import os

    try:
        if os.path.exists(OUTPUT_FILENAME):
            os.remove(OUTPUT_FILENAME) # one file at a time
            print('deleted old output')
    except Exception as e:
        print(e)

#####################    CRAWLER   #####################

class EbaySpiderSpider(scrapy.Spider):
    name = 'ebay_tester'
    
    #delete old scrapper_ouput.json
    delete_old_output()

    def start_requests(self):

        custom_settings = {
            'USER_AGENT': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/71.0.3578.80 Chrome/71.0.3578.80 Safari/537.36'
        }

        # allowed_domains = ['www.ebay.es']

        data_entries = get_queries()
        ebay_id_list = get_ebay_id_list()

        for entry in data_entries:
            # print(entry)

            query_title=    entry.get('query_title')
            
            if query_title == None: 
                continue
            
            query_title = query_title.lower()
            query_prod_state=     entry.get('query_prod_state')
            query_attribute_1=  entry.get('query_attribute_1')
            query_attribute_2=  entry.get('query_attribute_2')
            query_quantity= entry.get('query_quantity')
            target_category=entry.get('target_category')
            query_price=    entry.get('query_price')
            available_colors=   entry.get('available_colors')
            query_model=    entry.get('query_model')

            #query url might include 'reacondicionado' but query_title don't, so it will appear in serch but it isn't mandatory to filter= prod_titles without 'reacondicionado' will pass the filter
            query_url = create_url(query_title, query_prod_state)
            logging.info(f'this is the query url: {query_url}')
            # query_url = 'https://www.ebay.es/sch/15032/i.html?_from=R40&_nkw=iphone+11&LH_TitleDesc=0'

            yield scrapy.Request(url=query_url, callback=self.serp, meta={
                'start_url':query_url,
                'query_title':query_title, 
                'query_quantity':query_quantity, 
                'query_price':query_price,
                'ebay_id_list':ebay_id_list, 
                'available_colors':available_colors, 
                'target_category':target_category,
                'query_attribute_1':query_attribute_1,
                'query_attribute_2':query_attribute_2,
                'query_model':query_model,
                'query_prod_state':query_prod_state
                })

    #ebay serp with prod links
    def serp(self, response):

        def filter_price_title(serp_prods, query_title, query_price, ebay_id_list):

            #generate excluded kws, returns a dict with a list of kws to avoid. Iphone 11 != Iphone 11 pro 
            def get_excluded_kws(target_title):
                import traceback

                # given a list of kw
                # for kw in excluded_kws
                    # if dkw in query:
                        # append to present_kws
                # if there are some present_kws:
                    # remove those from exluded_kws
                
                # now you have a dict with 2 lists: excluded and present

                target_title = str(target_title)
                try:
                    excluded_kws = ['5g', 'pro', 'max', 'lite', 'ultra', 'plus', 'air','mini', 'active', '+']
                    present_kws  = []
                    for kw in excluded_kws:
                        if kw in target_title:
                            present_kws.append(kw)
                    
                    if len(present_kws) > 0:
                        for kw in present_kws:
                            excluded_kws.remove(kw)
                    try:
                        _dict = {'present_kw':present_kws, 'excluded_kws':excluded_kws}
                        return _dict
                    except:
                        _dict = {'excluded_kws':excluded_kws}
                        return _dict
                except Exception as e:
                    print(e)
                    traceback.print_exc()
                    return 'error'


            #used in filter to exclude serp_titles with excluded wk
            def excluded_kw_absence(serp_title, excluded_kws):
                '''filter excluding kws, iphone 12 pro != iphone 12 // if excluded kw in serp_title: continue'''

                # exc kws is a dict that contain included_kws and exc_kws, or only excluded_kws
                # this extracts the list of exc_kws from the dict
                excluded_kws = excluded_kws['excluded_kws']
                
                flag = 0
                for kw in excluded_kws:
                    if kw in serp_title.lower():
                        print(f'excluded kw: <{kw}> in <{serp_title}>')
                        flag +=1 # exc kw found in serp_title

                if flag == 0 :
                    print(f'this title pass {serp_title}')
                    return True # any exc kw in title
                elif flag != 0:
                    return False #some exc kw in title

            # gets the int price of the ebay product SERP
            def get_price(prod):

                serp_price = prod.xpath('.//span[@class="s-item__price"]/text()').get()

                # if that xpath doesn't work, try another
                if serp_price == None:
                    serp_price = prod.xpath('.//span[@class="ITALIC"]/text()').get()

                serp_price = serp_price.replace('$','').replace('€','')

                # from 1.248,55 to 1248
                if '.' in serp_price and ',' in serp_price:
                    serp_price = serp_price.split(',')[0]
                    serp_price = serp_price.replace('.','')
                    # print(f'detected , and . new price: {serp_price}')

                if ',' in serp_price: #677,80
                    serp_price = int(serp_price.split(',')[0]) #from "752,95 EUR" to 752
                elif '.' in serp_price: #1.250,45
                    serp_price = int(serp_price.replace('.', ',').split(',')[0])
                
                serp_price = int(serp_price)

                return serp_price

            #filter by title, split query in words, if all words in serp_title and any excluded kws =  yield the serp_link 
            def title_filter(query_title, serp_title, serp_link):
                s = query_title.split(' ') #split in words
                n = len(s)

                flag = 'present'
                # for each word of query, check if it's present in serp_title
                for i in range(n):
                    if s[i] not in serp_title:
                        flag = 'kw not present'
                
                if flag == 'present':
                    r = excluded_kw_absence(serp_title, excluded_kws)
                    if r:
                        print(f'filter passed: {serp_title}\nquery: {query_title}\n------------------------------')
                        return serp_link
                else:
                    print(f'this title NOT pass the filter:\n{serp_title}\nquery_title: {query_title}\n-------------------------')

            
            ################ End func declarations
            errors_n = 0
            url_list = []
            
            target_model = response.meta['query_model']
            ebay_id_list = response.meta['ebay_id_list']
            
            excluded_kws = get_excluded_kws(target_model)
            print(f'this are excluded kws: {excluded_kws}')

            number_bad_serp_price = 0 #used to count how many times cralwer doesn't get well serp_price or title, or other
            number_bad_serp_title = 0 #used to count how many times cralwer doesn't get well serp_price or title, or other
            number_bad_serp_link  = 0

            for prod in serp_prods:
                try:
                    prod_text = prod.get()
                    prod_text = prod_text.encode('utf-8')
                    serp_link = prod.xpath('.//div[@class="s-item__image"]/a[@tabindex="-1"]/@href').get() #all the same ??
                    serp_id = serp_link.split('itm/')[1].split('?')[0]
                    serp_title = prod.xpath('.//h3[@class="s-item__title"]/text()').get()
                    serp_title = serp_title.lower()
                    prod_state = prod.xpath('.//span[@class="SECONDARY_INFO"]').get()
                    
                    # avoid error with decoding titles
                    try:
                        print('\n serp_title: ', serp_title, '\n')
                    except UnicodeEncodeError:
                        print('continue!!--------')
                        continue

                    # serp_price = prod.xpath('.//span[@class="s-item__price"]/text()').get()
                    serp_price = get_price(prod)

                    #avoid broken items
                    if 'solo piezas' in prod_state:
                        continue
                    
                    #avoid breaking with problems getting serp data
                    if serp_price == None:
                        number_bad_serp_price += 1
                        print(f'bad serps price: {number_bad_serp_price}')
                        continue
                    elif number_bad_serp_title == None:
                        number_bad_serp_title += 1
                        print(f'bad serps title: {number_bad_serp_title}')
                        continue
                    elif number_bad_serp_link == None:
                        number_bad_serp_link += 1
                        print(f'bad serps title: {number_bad_serp_link}')
                        continue
                    
                except:
                    traceback.print_exc()
                    continue
                
                
                #if the prod is already in ebay_id = prod exist in prod_db -> ignore it
                if serp_id in ebay_id_list:
                    print(f'this ebay_id already exist {serp_id}')
                    continue
                #if the prod is cheaper than the filter price, ignore that prod
                if serp_price < query_price:
                    print(f'this price <{serp_price}> is too low for the filter price: <{query_price}>')
                    continue
                
                ##### begins title filter #####
                #filter by title, split query in words, if all words in serp_title and any excluded kws =  yield the serp_link 
                filtered_link = title_filter(query_title, serp_title, serp_link)
                if filtered_link:
                    url_list.append(filtered_link)
            
            return url_list

        ### end title filter ###
        start_url =    response.meta["start_url"]
        ebay_id_list = response.meta['ebay_id_list']
        query_title = response.meta['query_title']
        query_price = response.meta['query_price']
        target_category=response.meta['target_category']
        query_attribute_1= response.meta['query_attribute_1']
        query_attribute_2= response.meta['query_attribute_2']
        query_model     = response.meta['query_model']
        query_prod_state = response.meta['query_prod_state']
        available_colors = response.meta['available_colors']

        # print('in serp():', '\n',
        #                 'start_url',start_url,'\n',
        #                 'ebay_id_list',ebay_id_list,'\n',
        #                 'query_category',query_category,'\n',
        #                 'query_title',query_title,'\n',
        #                 'query_price',query_price,'\n',
        #                 'available_colors', available_colors, '\n'
        #                 # 'excluded_kws', excluded_kws,'\n'
        #                 )

        #get product  webelements from ebay serps
        serp_prods = response.xpath('//li[@class="s-item s-item__pl-on-bottom"]')
        print('-------serp prods count', len(serp_prods))

        #return url list of products that pass title and price
        filtered_prods = filter_price_title(serp_prods, query_title, query_price, ebay_id_list)
        
        #for some reason there're duplicateds, set the list
        filtered_prods = set(filtered_prods)
        print('filtered_prods: ',len(filtered_prods))

        for prod_url in filtered_prods:
            yield scrapy.Request(url= prod_url, callback=self.parse, meta={
                'start_url':start_url,
                'query':query_title, 
                'target_category':target_category,
                'query_attribute_1':query_attribute_1,
                'query_attribute_2':query_attribute_2,
                'query_model':query_model,
                'query_prod_state':query_prod_state,
                'available_colors':available_colors,
                })


    def parse(self, response):        

        #alternative func to get the specs (used as backup solution for the current func)
        def get_specs_with_bs(response):
            from bs4 import BeautifulSoup as bs

            # get the big box html with all specs
            # use bs to get the text using special separator &&
            #discriminate using "&& &&" replacing the separator to separate lines of the specs

            all_specs_box_div = response.xpath('//div[@class="ux-layout-section__item ux-layout-section__item--table-view"]').extract_first()

            soup = bs(all_specs_box_div)
            specs_bs_separated = soup.get_text(separator="&&")

            specs = specs_bs_separated.replace('&& &&','\n').replace('&&',' ').replace('  ','')

            return specs
        
        
        #this doesn't work well, in vscode viewer the csv looks fine
        #in excel csv viewer looks with bugs, strange breaklines
        def get_related_links(response):
            related_prod_list = []
            related_links = response.xpath('//li[@class="rtxt"]/parent::ul/li/a/@href').getall()#.extract_first()
            # my_print(len(related_prod_list))
            # for prod in related_links:
            #     related_prod_list.append(prod)
            return related_links

        # if the 1º used xpath fails, use a function to try all possible xpaths
        def get_price(price):
            
            price = response.xpath('//span[@id="convbidPrice"]/text()').get()

            if not price:        
                price = response.xpath('//span[@id="prcIsum"]/text()').get() # when a product is for direct selling and auction at the same time
            if price == None:           #notice the id is conviD and conviN, the're differents
                price = response.xpath('//span[@id="convbinPrice"]/text()').get()
            if price == None:   #this is spanish EUR price if prod is from spain
                price = response.xpath('//span[@class="notranslate"]/text()').extract_first()
            if price == None:
                price = response.xpath('//span[@id="mm-saleDscPrc"]/text()').get()
            if price == None:
                price = response.xpath('//span[@class="notranslate"]/text()').get()
            if price == None:
                price = response.xpath('//span[@class="notranslate "]/text()').get()
            if price == None:
                price = response.xpath('//span[@id="mm-saleDscPrc"]/text()').get()
            return price

        def get_shipping_price(response):
            # the order is important, sometimes they give the price in $ and other in €, is important to try to get 1º the €
            # try one xpath, if fails: try another

            # this has to go earlier than the next one because is the price in EUR. Else it will take price in USD or GBP
            shipping_price = response.xpath('//div[@class="ux-labels-values col-12 ux-labels-values--shipping"]//span[@class="ux-textspans ux-textspans--SECONDARY ux-textspans--BOLD"]/text()').get()

            if shipping_price == None:
                shipping_price = response.xpath('//div[@class="ux-labels-values col-12 ux-labels-values--shipping"]//span[@class="ux-textspans ux-textspans--BOLD"]/text()').get()
            if shipping_price == None:
                shipping_price = response.xpath('//span[@class="ux-textspans ux-textspans--BOLD"]//div[@class="ux-labels-values__values-content"]/text()').get()
            if shipping_price == None:
                shipping_price = response.xpath('//span[@class="ux-textspans ux-textspans--POSITIVE ux-textspans--BOLD"]/text()').extract_first()
            if shipping_price == None: #recogida local
                shipping_price = response.xpath('//div[@class="ux-labels-values col-12 ux-labels-values__column-last-row ux-labels-values--localPickup"]')
                if shipping_price: 
                    return 'local pick up'

            # include? sif EUR in shipping or gratis in shipping: return
            return shipping_price

        #this pics are thumbnails, in filter I change the URL pattern to access the big pics
        def get_ebay_pics(pics_selector):
            pics_list = []
            lis = pics_selector.xpath('.//li')
            for li in lis:
                pic = li.xpath('.//img/@src').get()
                pics_list.append(pic)
            return pics_list

        # product returns options: 'Devoluciones 30 días'
        def get_returns(response):
            from bs4 import BeautifulSoup as bs

            #get html
            #with bs, get text from html 
            #remove unwanted part of text
            
            returns_html = response.xpath('//div[@class="ux-layout-section ux-layout-section--returns"]').extract_first()
            
            soup = bs(returns_html)
            returns_text = soup.get_text()
            returns_text = returns_text.replace('| El vendedor paga el envío de la devolución | Ver detalles- Más información sobre devoluciones', '')

            return returns_text

        # sometimes ebay retrieves prods that are already soldout
        def check_if_prod_soldout(response):

            product_sold_out_text = response.xpath('//span[contains(text(),"Este artículo está agotado")]')
            # if that xpath didn't work, try this one
            if not product_sold_out_text:
                product_sold_out_text = response.xpath('//span[contains(text(),"El vendedor ha finalizado este anuncio")]')
            
            return product_sold_out_text

        # Used to get warranty and or state, in ebay just below the title
        def get_subtitle(response):
            subtitle = response.xpath('//span[@class="ux-textspans ux-textspans--ITALIC"]/text()').get()
            
            if not subtitle:
                subtitle = response.xpath('//span[@class="topItmCndDscMsg"]/text()').get()
            if not subtitle:
                subtitle = response.xpath('//div[@id="subTitle"]/text()').get()
            
            if subtitle:
                return subtitle
            

        #################   PARSE FUNCTION   ####################
        title = response.xpath('//h1[@class="x-item-title__mainTitle"]/span/text()').get()
        query = response.meta["query"]
        target_category = response.meta['target_category']
        query_attribute_1= response.meta['query_attribute_1']
        query_attribute_2= response.meta['query_attribute_2']
        query_model      = response.meta['query_model']
        query_prod_state = response.meta['query_prod_state']
        available_colors = response.meta['available_colors']
        
        variable_prod    = response.xpath('//span[@id="sel-msku-variation"]').get()
        ebay_article_id  = response.xpath('//div[@id="descItemNumber"]/text()').extract_first()
        prod_url    = response.url
        ebay_vendor = response.xpath('//span[@class="ux-textspans ux-textspans--PSEUDOLINK ux-textspans--BOLD"]/text()').get()
        product_state = response.xpath('//div[@class="d-item-condition-text"]//span/text()').get()
        #this takes the raw html, in filter is processed to give the number of days
        shipping_time = response.xpath('//div[@class="ux-labels-values col-12 ux-labels-values__column-last-row ux-labels-values--deliverto"]').extract_first()

        #if the prod is not from Spain, the xpath for shippment changes
        # when prod is not in spain, this identifies if its shipped to spain or not
        served_area = response.xpath('//span[@itemprop="areaServed"]/text()').get()
        reviews = response.xpath('//div[@class="reviews"]/text()').extract_first()
        seller_votes = response.xpath('//div[@class="ux-seller-section__item--seller"]//a/span[@class="ux-textspans ux-textspans--PSEUDOLINK"]/text()').get()
        payment_methods = pay = response.xpath('//div[@class="ux-labels-values__values-content"]//span/@aria-label').getall()
        #translated later in filter.py
        prod_specs_html = response.xpath('//div[@class="ux-layout-section__row"]')
        import_taxes = response.xpath('//div[@class="ux-labels-values col-12 ux-labels-values--importCharges"]//span[@class="ux-textspans ux-textspans--BOLD"]/text()').get()

        product_sold_out_text = check_if_prod_soldout(response)
        prod_specs = get_specs_with_bs(prod_specs_html)
        subtitle   = get_subtitle(response)
        returns    = get_returns(response)
        shipping_price = get_shipping_price(response)

        #this pic is always available, but not always there are more than one. Use try: to find other pics, if any pics= main
        ebay_main_pic_url = response.xpath('//img[@id="icImg"]/@src').get()
        try: 
            #this pics are thumbnails, later in filter.py I change the url pattern to access the big pictures using the thumbnail URL
            pics_selector = response.xpath('//ul[@class="lst icon"]')[1]
            ebay_pics  = get_ebay_pics(pics_selector)
            ebay_pics.append(ebay_main_pic_url)
        except Exception as e:
            ebay_pics = ebay_main_pic_url
            print(e)

        try:
            reviews = reviews.replace('\n','')
        except:
            pass

        try:
            served_area = served_area.replace('\n','')
        except:
            pass

        
        # article location: in desuse, accepting all results
        #depending on article location shows on price or another
        # article_location = response.xpath('//span[@itemprop="availableAtOrFrom"]/text()').get()
        # if 'España' not in article_location: # or EU not in article location
        #     pass

        #many prods are from UK, they show price in GBP and EUR,
        # this tries first to take EUR prices with prods with both options
        price = response.xpath('//span[@id="convbinPrice"]/text()').get()
        if price == None or '':
            price = get_price(price) # there're a lot of alternatives, so I made a function

        #this is the iframe with prod_description inside, to get it it's needed another request to get the info inside the iframe
        #scrape all the data and make another request to the iframe's url and pass all the scrapped data as meta.
        #some prods don't have iframe. if iframe == none take the description and yield to finish this item
        iframe_description_url = response.xpath('//div[@id="desc_wrapper_ctr"]/div/iframe/@src').get()
        if iframe_description_url == None:
            iframe_description_url = 'not present' # set it this way to yield it in the output
        else: #if iframe url exist, make a request to it
            # now call the iframe parser and give it all the info inside meta
            yield Request(url=iframe_description_url, callback=self.iframe,
                        meta={
                            'title':title,
                            'price':price, 
                            'query':query,
                            'shipping_time':shipping_time,
                            'variable_prod':variable_prod,
                            'returns':returns,
                            'shipping_price':shipping_price,
                            'ebay_article_id':ebay_article_id,
                            'prod_url':prod_url,
                            'ebay_vendor':ebay_vendor,
                            'seller_votes':seller_votes,
                            'payment_methods':payment_methods,
                            'product_state':product_state,
                            'product_sold_out_text':product_sold_out_text,
                            'served_area':served_area,
                            'reviews':reviews,
                            'prod_specs':prod_specs,
                            'import_taxes':import_taxes,
                            'target_category':target_category,
                            'query_attribute_1':query_attribute_1,
                            'query_attribute_2':query_attribute_2,
                            'query_model':query_model,
                            'query_prod_state':query_prod_state,
                            'ebay_pics':ebay_pics,
                            'available_colors':available_colors,
                            'subtitle':subtitle,
                            'iframe_description_url':iframe_description_url
                            })


    #this scrapes the prod description in iframe
    def iframe(self,response):
        
        prod_description = response.xpath('//div[@id="ds_div"]').extract()
        if prod_description == None:
            prod_description = 'not present'

        #get all data from meta
        title   =   response.meta['title']
        price   =   response.meta['price']
        query   =   response.meta['query']
        shipping_time = response.meta['shipping_time']
        variable_prod = response.meta['variable_prod']
        returns =   response.meta['returns']
        shipping_price=response.meta['shipping_price']
        ebay_article_id=response.meta['ebay_article_id']
        prod_url     =  response.meta['prod_url']
        ebay_vendor  =  response.meta['ebay_vendor']
        seller_votes =  response.meta['seller_votes']
        payment_methods=response.meta['payment_methods']
        product_state  =response.meta['product_state']
        prod_specs   =  response.meta['prod_specs']
        served_area  =  response.meta['served_area']
        reviews  =   response.meta['reviews']
        product_sold_out_text = response.meta['product_sold_out_text']
        import_taxes     =  response.meta['import_taxes']
        target_category  =  response.meta['target_category']
        query_attribute_1=  response.meta['query_attribute_1']
        query_attribute_2=  response.meta['query_attribute_2']
        query_model      =  response.meta['query_model']
        query_prod_state =  response.meta['query_prod_state']
        ebay_pics        =  response.meta['ebay_pics']
        available_colors =  response.meta['available_colors']
        subtitle         =  response.meta['subtitle']
        iframe_description_url = response.meta['iframe_description_url']
        
        yield {'title':title,'price':price, 'query':query,
        'shipping_time':shipping_time,     'variable_prod':variable_prod,
        'returns':returns,                'shipping_price':shipping_price,
        'ebay_article_id':ebay_article_id,'prod_url':prod_url,
        'ebay_vendor':ebay_vendor,      'seller_votes':seller_votes,
        'payment_methods':payment_methods,
        'prod_specs':prod_specs,        'product_state':product_state, 
        'prod_description':prod_description,
        'served_area':served_area,     
        'product_sold_out_text':product_sold_out_text,
        'import_taxes':import_taxes, 
        'query_attribute_1':query_attribute_1,
        'query_attribute_2':query_attribute_2,
        'query_model':query_model,
        'query_prod_state':query_prod_state,
        'ebay_pics':ebay_pics,
        'available_colors':available_colors,
        'subtitle':subtitle,
        'iframe_description_url':iframe_description_url,
        'target_category':target_category,
        # 'reviews':reviews,
         }


 # item.add_xpath('warranty', '')
        # item.add_xpath('num_availables','')
        # item.add_xpath('','')

        # item.add_xpath('score', './/div[@class="kVNDLtqL"]/span/text()',
        # MapCompose(self.quitarDolar))
        # # Utilizo Map Compose con funciones anonimas
        # item.add_xpath('descripcion', '//div[@class="ui_column  "]//div[@class="cPQsENeY"]//text()', # //text() nos permite obtener el texto de todos los hijos
        # MapCompose(lambda i: i.replace('\n', '').replace('\r', '')))
