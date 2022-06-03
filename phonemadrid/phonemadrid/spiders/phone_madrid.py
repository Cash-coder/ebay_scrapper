# scrapy crawl phone_m_spider -o phone_m_output.json
# WITH_MOVE (scrapy crawl phone_m_spider -o phone_m_output.json) -and (move phone_m_output.json "C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder" -force)

from venv import create
import scrapy
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from time import sleep
import logging


OUTPUT_FILENAME = r'C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\ebay\phonemadrid\phonemadrid\spiders\phone_m_output.json'
PRODUCTS_DB     = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\PRODS_DB.xlsx"
GAPS_FILE       = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\gaps_file.xlsx"

# IN GAPS_FILE
GAPS_TITLE_COL  = 0
GAPS_PROD_STATE = 1
GAPS_MODEL_COL  = 2
GAPS_ATTR_1_COL = 3
GAPS_ATTR_2_COL = 4
GAPS_QUANTITY   = 5
GAPS_CATEGORY   = 6
GAPS_MEAN_PRICE = 7
GAPS_MIN_PRICE  = 8
PRODSDB_EBAY_ID = 'R'

# get data to parse
# create url, search that url
# in serps:
    # identify prods
    # filter prods by title and id
    # if pass filters add to list the url's

    # if there's another page, go to another page and repeat
    # if no more pages
        # for url in filter_passed_urls
            # parse

def get_id_list():
    from openpyxl import load_workbook

    wb = load_workbook(PRODUCTS_DB)
    ws = wb.active

    ebay_ids = []

    for cell in ws[PRODSDB_EBAY_ID]:
        if cell.value != None:
            # print(cell.value)
            ebay_ids.append(cell.value)
    print(f'ebay id len: {len(ebay_ids)}')
    return ebay_ids

def get_queries():
    """ this takes a file named queries.csv and creates url's to search
    by the scrapper """
    from openpyxl import load_workbook

    #gaps_file
    wb = load_workbook(GAPS_FILE)
    ws = wb.active

    #get all the rows
    for row in ws.iter_rows(values_only=True, min_row=3):
        #query_title = iphone 12 // later I add: <iphone 12 64 gb verde> // and finally I create a different var query_search_title = query_title + prod_state// this is used to include "reacondicionado" in query, but not like mandatory to pass the filter
        query_title       = row[GAPS_TITLE_COL]
        query_prod_state  = row[GAPS_PROD_STATE]
        query_model       = row[GAPS_MODEL_COL]
        query_attribute_1 = row[GAPS_ATTR_1_COL] #FI: GB
        query_attribute_2 = row[GAPS_ATTR_2_COL] #FI: color
        query_quantity    = row[GAPS_QUANTITY]
        target_category   = row[GAPS_CATEGORY]
        query_price       = row[GAPS_MIN_PRICE] #filter price, like min 300€, not prods of 5€
        mean_price        = row[GAPS_MEAN_PRICE] # FI: iphone 12 != iphone 12 PRO // exclue those that has MAX or PRO in title

        if query_title == None : continue
        
        entry = {
        'query_title':query_title,
        'query_attribute_1':query_attribute_1,
        'query_attribute_2':query_attribute_2,
        'query_quantity':query_quantity,
        'target_category':target_category,
        'query_price':query_price,
        'mean_price':mean_price,
        'query_model':query_model,
        'query_prod_state':query_prod_state
        }

        yield entry


def delete_old_output():
    import os

    try:
        if os.path.exists(OUTPUT_FILENAME):
            os.remove(OUTPUT_FILENAME) # one file at a time
            print('deleted old output')
        else:
            print('file path to delete old output does not exist!')
    except Exception as e:
        print(e)

def create_url(query_title):

    formatted_query = query_title.replace(' ', '+')
   
    url  = f'https://www.phonemadrid.com/es/buscar?controller=search&s={formatted_query}'
    
    print(f'----- gap url: {url}')

    return url


class PhoneMadridSpider(scrapy.Spider):

    name = 'phone_m_spider'

    delete_old_output()


    def start_requests(self):

        # query_title     = 'HP ProBook 6550b Intel Core i5 Windows7'
        # target_category = 'smartphones'
        # ids_list        = ['aasdas','195-774']
        
        data_queries = get_queries()

        for entry in data_queries:
            print(entry)

            query_title=    entry.get('query_title')
            
            if query_title == None: 
                continue
            
            query_title       = query_title.lower()
            query_attribute_1 = entry.get('query_attribute_1')
            query_attribute_2 = entry.get('query_attribute_2')
            target_category   = entry.get('target_category')
            query_model       = entry.get('query_model')

            query_url = create_url(query_title)
            ids_list  = get_id_list()
            print(f'query_url: {query_url}\n')

            yield scrapy.Request(url=query_url, callback=self.pagination, meta={
                'query_title':       query_title, 
                'target_category':   target_category,
                'ids_list':          ids_list,
                'query_attribute_1': query_attribute_1,
                'query_attribute_2': query_attribute_2,
                'query_model':       query_model
                })

    
    def pagination(self, response):
        
        # def filters
        # get prods
        # filter by id
        # filter by title
        # pass filtered to parse
        # if next button, use pagination again

        def filter_by_id(id_list, serp_prods):
            
            # filtered_prods_by_id = []

            # for prod in serps_prods
                # get prod_id
                # if id in ids_list: continue
                # append + return  prod

            filtered_prods_by_id = []
            for prod in serp_prods:
                prod_id   = prod.xpath('./@data-id-product').get()
                prod_id_attribute   = prod.xpath('./@data-id-product-attribute').get()

                prod_complete_id = f'{prod_id}-{prod_id_attribute}'
                
                if prod_complete_id in id_list:
                    print(f'\nthis prod_id already exist {prod_complete_id}\n')
                    continue                            

                filtered_prods_by_id.append(prod)
            
            return filtered_prods_by_id
            # End filter_by_id()

        # begin def filter_by_title()
        def filter_by_title(filtered_prods_by_id, query_title):

            # def filter_by_id()
            # def get_exc_kws()
            # def exc_kws_absence()

            # exc_kws = get_exc_kws()

            # for prod in filtered_by_id:
                # prod_title  = prod.xpath()
                # kws_are_absent = excluded_kw_absence(serp_title, excluded_kws):
                # if kws_are_absent:
                    # yield prod

            
            # START DEFINITIONS
            #generate excluded kws, returns a dict with a list of kws to avoid. Iphone 11 != Iphone 11 pro 
            def get_excluded_kws(target_title):
                import traceback

                # given a list of kw
                # for kw in excluded_kws
                    # if dkw in query:
                        # append to present_kws
                # if there are some present_kws:
                    # remove those from exluded_kws
                
                # now you have a dict with 2 lists: excluded and present

                target_title = str(target_title).lower()
                try:
                    excluded_kws = [' 5g ', ' pro ', ' max ', ' lite ', ' ultra ', ' plus ', ' air ',' mini ', ' active ', '+', ' xr ']
                    present_kws  = []
                    for kw in excluded_kws:
                        if kw in target_title:
                            present_kws.append(kw)
                    
                    if len(present_kws) > 0:
                        for kw in present_kws:
                            excluded_kws.remove(kw)
                    
                    print(f'title: {target_title}\nexc_kws {excluded_kws}\npresent: {present_kws}')

                    try:
                        _dict = {'present_kw':present_kws, 'excluded_kws':excluded_kws}
                        return _dict
                    except:
                        _dict = {'excluded_kws':excluded_kws}
                        return _dict
                except Exception as e:
                    print(e)
                    traceback.print_exc()
                    return 'error'


            #used in filter to exclude serp_titles with excluded wk
            def excluded_kw_absence(serp_title, excluded_kws):
                '''filter excluding kws, iphone 12 pro != iphone 12 // if excluded kw in serp_title: continue'''
                
                serp_title = serp_title.lower()

                # exc kws is a dict that contain included_kws and exc_kws, or only excluded_kws
                # this extracts the list of exc_kws from the dict
                excluded_kws = excluded_kws['excluded_kws']
                
                flag = 0
                for kw in excluded_kws:
                    if kw in serp_title:
                        print(f'excluded kw: <{kw}> in <{serp_title}>')
                        flag +=1 # exc kw found in serp_title

                if flag == 0 :
                    print(f'this title pass {serp_title} | {exc_kws}')
                    return True # any exc kw in title
                elif flag != 0:
                    return False #some exc kw in title
                # END DEFINITION excluded_kw_absence
            
            def check_all_target_kws_present(target_title, prod_title):
                target_kws = target_title.lower().split(' ')

                prod_title = prod_title.lower()

                for kw in target_kws:
                    if kw not in prod_title:
                        print(f'--------------missing kw "{kw}" from target:{query_title} prod_title: {prod_title}')
                        return False
                
                return True

            # filter_by_title() code

            filtered_prods_by_title = []

            exc_kws = get_excluded_kws(query_title)
            print(f'exc_kws are: {exc_kws}')

            for prod in filtered_prods_by_id:
                prod_title = prod.xpath('.//div[@class="product-container"]/div[@class="second-block"]/h5[@class="product-name"]/a/@title').get()

                all_target_kws_are_present = check_all_target_kws_present(query_title, prod_title)
                if not all_target_kws_are_present:
                    print(f'missing kw, target {query_title}, {prod_title}')
                    continue

                kws_are_absent = excluded_kw_absence(prod_title, exc_kws)

                if kws_are_absent:
                    link =  prod.xpath('.//a[@class="product-cover-link"]/@href').get()
                    filtered_prods_by_title.append(link)

            return filtered_prods_by_title
        # END def filter_by_title() 

        target_category   = response.meta['target_category']
        query_title       = response.meta['query_title']
        ids_list          = response.meta['ids_list']
        query_model       = response.meta['query_model']
        query_attribute_1 = response.meta['query_attribute_1']
        query_attribute_2 = response.meta['query_attribute_2']

        prods = response.xpath('//article')

        prods_filtered_by_id = filter_by_id(ids_list, prods)
        print(f'filtered by id: {prods_filtered_by_id}')

        filtered_links    = filter_by_title(prods_filtered_by_id, query_title)
        print(f'filtered_links {filtered_links}')


        for link in filtered_links: 
            yield scrapy.Request(url=link, callback=self.parse, meta={
                'url':               link,
                'query_title':       query_title,
                'target_category':   target_category,
                'ids_list':          ids_list,
                'query_attribute_1': query_attribute_1,
                'query_attribute_2': query_attribute_2,
                'query_model':       query_model,
            } )


        next_page = response.xpath('//a[@class="next js-search-link"]/@href').get()
        if next_page:
            yield scrapy.Request(url=next_page, callback=self.pagination, meta={
                'query_title':query_title,
                'ids_list':query_title,
                'target_category':target_category,
                'query_attribute_1': query_attribute_1,
                'query_attribute_2': query_attribute_2,
                'query_model':       query_model
                })


    def parse(self, response):
        
        prod_url          = response.meta['url']
        target_category   = response.meta['target_category']
        query_model       = response.meta['query_model']
        query_attribute_1 = response.meta['query_attribute_1']
        query_attribute_2 = response.meta['query_attribute_2']

        button_disabled = response.xpath('//button[@class="btn add-to-cart"]/@disabled').get()

        # if button disabled -> prod no available
        if not button_disabled:

            price = response.xpath('//span[@class="price product-price"]/text()').get()
            title = response.xpath('//h1/text()').get()
            description = response.xpath('//div[@id="tabPhoneContent"]').extract()

            pics         = response.xpath('//li[@class="thumb-container"]/a/@data-zoom-image').getall()
            prod_id      = response.xpath('//div[@class="ed_item "]/@data-id-product').get()
            prod_id_attr = response.xpath('//div[@class="ed_item "]/@data-id-product-attribute').get()
            complete_id = f'{prod_id}-{prod_id_attr}'

            yield {
                'description': description,
                'price': price,
                'title': title,
                'pics' : pics,
                'prod_url':    prod_url,
                'complete_id': complete_id,
                'target_category':   target_category,
                'query_attribute_1': query_attribute_1,
                'query_attribute_2': query_attribute_2,
                'query_model':       query_model
            }
