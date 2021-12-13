# cd "C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\ebay\ebay_scrapper & API\ebay_scrapper_api\ebay_scrapper_api\spiders"
# scrapy crawl ebay-spider -o crawler_output.json
# scrapy crawl ebay-spider -o output.csv -t csv -a CSV_DELIMITER="|"
''' output file has to be named ebay-spider-output.json '''
import traceback
import scrapy
import csv
from scrapy.http.request import Request
from scrapy.spiders import Spider
from scrapy.selector import Selector
from scrapy.loader.processors import MapCompose
from scrapy.spiders import CrawlSpider, Rule
from scrapy.linkextractors import LinkExtractor
from bs4 import BeautifulSoup as bs4
from scrapy.http import HtmlResponse
import logging
from time import sleep

# logging.basicConfig(level=logging.DEBUG)
# logging.basicConfig(level=logging.INFO)
# logging.basicConfig(level=logging.WARNING)

#for my_print()
from colorama import init
from termcolor import colored

init()

GAPS_FILE = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\gaps_file.xlsx"
OUTPUT_FILENAME = 'crawler_output.json' # used only to delete old output, not to create the new one. New creates from cmd when launching spider

# IN GAPS_FILE
GAPS_TITLE_COL  = 0
GAPS_PROD_STATE = 1
GAPS_MODEL_COL  = 2
GAPS_ATTR_1_COL = 3
GAPS_ATTR_2_COL = 4
GAPS_QUANTITY   = 5
GAPS_CATEGORY   = 6
GAPS_AVAILABLE_COLORS= 7
GAPS_MIN_PRICE  = 8
GAPS_EBAY_ID    = 'J2' #they're all in the same cell


def my_print(text, color='green', mode='normal',**id):
    ''' red,green,yellow,blue // mode="lines" to print a list line by line in BLUE"'''

    if mode == 'normal':
        if color == 'red':
            print(colored(id, 'white','on_red'))
            print(colored(text, 'white','on_red'))
        elif color == 'green':
            print(colored(id, 'white','on_green'),sep="\n")
            print(colored(text, 'white','on_green'),sep="\n")
        elif color == 'yellow':
            print(colored(id, 'blue','on_yellow'))
            print(colored(text, 'blue','on_yellow'))
        elif color == 'blue':
            print(colored(id, 'white','on_blue'))
            print(colored(text, 'white','on_blue'))

    elif mode=='lines':
        for item in text:
            print(colored(item, 'white','on_blue'))

#used to avoid process products we already have
def  get_ebay_id_list():
    from openpyxl import load_workbook
    #gaps_file
    wb = load_workbook(GAPS_FILE)
    ws = wb.active

    try:
        ebay_id_list = ws[GAPS_EBAY_ID].value
        ebay_id_list = ebay_id_list.split(',')
        print(f'---ebay_id_list{ebay_id_list}')
        return ebay_id_list
    except Exception as e:
        print(f'in get_ebay_id_list(): {e}')
        traceback.print_exc()
        return None

def get_queries():
    """ this takes a file named queries.csv and creates url's to search
    by the scrapper """
    from openpyxl import load_workbook

    data_queries = []

    # global first_chunk
    # global last_chunk

    #gaps_file
    wb = load_workbook(GAPS_FILE)
    ws = wb.active

    #get the ebay_id_list
    ebay_id_list = get_ebay_id_list()

    #get all the rows
    for row in ws.iter_rows(values_only=True, min_row=3):
        #query_title = iphone 12 // later I add: <iphone 12 64 gb verde> // and finally I create a different var query_search_title = query_title + prod_state// this is used to include "reacondicionado" in query, but not like mandatory to pass the filter
        query_title =   row[GAPS_TITLE_COL]
        query_prod_state=row[GAPS_PROD_STATE]
        query_model =   row[GAPS_MODEL_COL]
        query_attribute_1=row[GAPS_ATTR_1_COL] #FI: GB
        query_attribute_2=row[GAPS_ATTR_2_COL] #FI: color
        query_quantity= row[GAPS_QUANTITY]
        target_category= row[GAPS_CATEGORY]
        query_price=    row[GAPS_MIN_PRICE] #filter price, like min 300€, not prods of 5€
        available_colors=   row[GAPS_AVAILABLE_COLORS] # FI: iphone 12 != iphone 12 PRO // exclue those that has MAX or PRO in title

        if query_title == None : continue
        
        print(
            'query_title: ',query_title, '\n',
            'query_model: ',query_model, '\n',
            'query_attribute_1: ',query_attribute_1, '\n',
            'query_attribute_2: ',query_attribute_2, '\n',
            'query_quantity: ',query_quantity, '\n',
            'target_category: ',target_category, '\n',
            'query_price: ',query_price, '\n',
            'available_colors: ',available_colors, '\n'
            'query_prod_state', query_prod_state
        )

        #OLD when I targeted for colors, now I target only for GB and apply colors later
        #if query attr (color) exists, append it to title: iphone -> iphone amarillo
        # if query_attribute_1 != None and query_attribute_1 != '':
        #     query_title = query_model + ' ' + query_attribute_1
        #     #if it's a tuple like negro,grafito. Take the 1º as attr and append it query_title and the 2º as var alt_attr_1
        #     if ',' in query_attribute_1:
        #         query_attribute_1 = query_attribute_1.split(',')
        #         query_attribute_1 = query_attribute_1[0]
        #         query_alt_attr_1 = query_attribute_1[1]
        #     else: #apply deafault values in case some attr dones't exist, to append to the list cleanly
        #         query_alt_attr_1 = 'default'

        # if query_attribute_2 != None and query_attribute_2 != '':
        #     query_title = query_title + ' ' + query_attribute_2
        #     if ',' in query_attribute_2:
        #         query_attribute_2 = query_attribute_2.split(',')
        #         query_attribute_2 = query_attribute_2[0]
        #         query_alt_attr_2 = query_attribute_2[1]
        #     else:
        #         query_alt_attr_2 = 'default'

        #NEW target only for GB if they're specified
        # if query_attribute_1 != None:
            # query_title = query_model + ' ' + query_attribute_1
        # else: #if attr1 not specified, search just the model
            # query_title = query_model

        #add data to list, return list
        # entry = [query_title, query_quantity, target_category, query_alt_attr_1, query_alt_attr_2, query_price, excluded_kws]
        entry = {
        'query_title':query_title,
        'query_attribute_1':query_attribute_1,
        'query_attribute_2':query_attribute_2,
        'query_quantity':query_quantity,
        'target_category':target_category,
        'query_price':query_price,
        'available_colors':available_colors,
        'query_model':query_model,
        'query_prod_state':query_prod_state
        }

        data_queries.append(entry)

    entry = [ebay_id_list, data_queries]
    return entry


def create_url(query_category, query_title, query_prod_state):
    # in gaps file is specified new or not new. Include some kws in the query to maximize results fo each
    if query_prod_state == 'not_new':
        query_title = query_title + ' usado'
    elif query_prod_state == 'new':
        query_title = query_title + ' nuevo'
    
    

    # NEW using search instead of picking categories
    first_chunk     = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
    formatted_query = query_title.replace(' ','+')
    second_chunk    = '&_sacat=0&rt=nc&LH_BIN=1'

    query_url = first_chunk + formatted_query + second_chunk
    print(f'------------- {query_url}')

    #WARNING! use url with 200 results and buy only option
    #use one or other chunks based on target category
    # if query_category == 'smartphones':
    #     first_url_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
    #     second_url_chunk = '&_sacat=9355&LH_TitleDesc=0&rt=nc&LH_BIN=1&_ipg=200'
    # elif query_category == 'smartwatches':
    #     first_url_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
    #     second_url_chunk = '&_sacat=178893&rt=nc&LH_BIN=1&_ipg=200'
    # elif query_category == 'laptops':
    #     first_url_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
    #     second_url_chunk = '&_sacat=175672&rt=nc&LH_BIN=1&_ipg=200'
    # elif query_category == 'auriculares':
    #     first_url_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
    #     second_url_chunk = '&_sacat=112529&rt=nc&LH_BIN=1&_ipg=200'
    # elif query_category == 'televisors':
    #     first_url_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
    #     second_url_chunk = '&_sacat=11071&rt=nc&LH_BIN=1&_ipg=200'
    # elif query_category == 'tablets':
    #     first_url_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
    #     second_url_chunk = '&_sacat=171485&rt=nc&LH_BIN=1&_ipg=200'
    # elif query_category == 'consolas' or 'videogames':
    #     first_url_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
    #     second_url_chunk = '&_sacat=139971&rt=nc&LH_BIN=1&_ipg=200'
    # elif query_category == 'digital cameras':
    #     first_url_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
    #     second_url_chunk = '&_sacat=31388&rt=nc&LH_BIN=1&_ipg=200'

    # #create the url
    # query_title = query_title.replace(' ', '+')
    # query_url = first_url_chunk + query_title + second_url_chunk

    return query_url

def delete_old_output():
    import os
    import shutil

    if os.path.exists(OUTPUT_FILENAME):
        os.remove(OUTPUT_FILENAME) # one file at a time
        print('deleted old output')


#####################              #####################
#####################    CRAWLER   #####################

# class EbaySpiderSpider(CrawlSpider):
class EbaySpiderSpider(scrapy.Spider):
    print('class')
    name = 'ebay-spider'
    
    #delete old scrapper_ouput.json
    delete_old_output()

    def start_requests(self):

        custom_settings = {
            'USER_AGENT': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/71.0.3578.80 Chrome/71.0.3578.80 Safari/537.36'
        }

        allowed_domains = ['www.ebay.es']

        data_entry = get_queries()
        data_queries = data_entry[1]
        ebay_id_list = data_entry[0]

        for entry in data_queries:
            print(entry)

            query_title=    entry.get('query_title')
            
            if query_title == None: 
                continue
            
            query_title = query_title.lower()
            query_prod_state=     entry.get('query_prod_state')
            query_attribute_1=  entry.get('query_attribute_1')
            query_attribute_2=  entry.get('query_attribute_2')
            query_quantity= entry.get('query_quantity')
            target_category=entry.get('target_category')
            query_price=    entry.get('query_price')
            available_colors=   entry.get('available_colors')
            query_model=    entry.get('query_model')

            # query_title = entry[0]
            # query_quantity = entry[1]
            # target_category = entry[2]
            # query_alt_attr_1 = entry[3] # FI: "grafito" instead of "negro"
            # query_alt_attr_2 = entry[4] # FI: "grafito" instead of "negro"
            # query_price = entry[5]
            # available_colors = entry[6]
            # available_colors = available_colors.split(',')
            
            #query url might include 'reacondicionado' but query_title don't, so it will appear in serch but it isn't mandatory to filter= prod_titles without 'reacondicionado' will pass the filter
            query_url = create_url(target_category, query_title, query_prod_state)
            logging.info(f'this is the query url: {query_url}')
            # query_url = 'https://www.ebay.es/sch/15032/i.html?_from=R40&_nkw=iphone+11&LH_TitleDesc=0'

            url_list = []
            print(f'declared url_list, {url_list}')
            
            yield scrapy.Request(url=query_url, callback=self.serp, meta={
                'start_url':query_url,
                'query_title':query_title, 
                'query_quantity':query_quantity, 
                'query_price':query_price,
                'ebay_id_list':ebay_id_list, 
                'available_colors':available_colors, 
                'target_category':target_category,
                'query_attribute_1':query_attribute_1,
                'query_attribute_2':query_attribute_2,
                'query_model':query_model,
                'url_list':url_list,
                'query_prod_state':query_prod_state
                })

    #ebay serp with prod links
    def serp(self, response):

        def filter_price_title(serp_prods, query_title, query_price, ebay_id_list):
            #for name, create excluded kws -> iphone 11 pro -> exc_kws = max, plus....

            #generate excluded kws, returns a dict with a list of kws to avoid. Iphone 11 != Iphone 11 pro 
            def get_excluded_kws(target_title):
                try:
                    excluded_kws = ['5g', 'pro', 'lite', 'ultra', 'plus', 'air','mini', 'active', '+']
                    for kw in excluded_kws:
                        if kw in target_title:
                            _index = excluded_kws.index(kw)
                            present_kw =  excluded_kws.pop(_index)
                    try:
                        _dict = {'present_kw':present_kw, 'excluded_kws':excluded_kws}
                        return _dict
                    except:
                        _dict = {'excluded_kws':excluded_kws}
                        return _dict
                except Exception as e:
                    print(e)
                    traceback.print_exc()
                    return 'error'

            #used in filter to exclude serp_titles with excluded wk
            def excluded_kw_absence(serp_title, excluded_kws):
                '''workaround to exclude kw, iphone 12 pro != iphone 12 // if excluded kw in serp_title: continue'''
                
                flag = 0
                for kw in excluded_kws:
                    if kw in serp_title.lower():
                        print(f'flag +1 kw: {kw} in <{serp_title}>')
                        flag +=1 # exc kw found in serp_title

                if flag == 0 :
                    print(f'this title pass {serp_title}')
                    return True # any exc kw in title
                elif flag != 0:
                    return False #some exc kw in title

            #Enf func declaration, begins code
            errors_n = 0
            url_list = response.meta['url_list']
            
            target_model = response.meta['query_model']
            excluded_kws = get_excluded_kws(target_model)

            number_bad_serp_price = 0 #used to count how many times cralwer doesn't get well serp_price or title, or other
            number_bad_serp_title = 0 #used to count how many times cralwer doesn't get well serp_price or title, or other
            number_bad_serp_link  = 0

            for prod in serp_prods:
                try:
                    prod_text = prod.get()
                    prod_text = prod_text.encode('utf-8')
                    serp_link = prod.xpath('.//div[@class="s-item__image"]/a[@tabindex="-1"]/@href').get() #all the same ??
                    serp_id = serp_link.split('itm/')[1].split('?')[0]
                    serp_title = prod.xpath('.//h3[@class="s-item__title"]/text()').get()
                    print('\n serp_title: ', serp_title, '\n')
                    serp_title = serp_title.lower()
                    serp_price = prod.xpath('.//span[@class="s-item__price"]/text()').get()
                    prod_state = prod.xpath('.//span[@class="SECONDARY_INFO"]').get()

                    #avoid broken items
                    if 'solo piezas' in prod_state:
                        continue
                    
                    #avoid breaking with problems getting serp data
                    if serp_price == None:
                        number_bad_serp_price += 1
                        print(f'bad serps price: {number_bad_serp_price}')
                        continue
                    elif number_bad_serp_title == None:
                        number_bad_serp_title += 1
                        print(f'bad serps title: {number_bad_serp_title}')
                        continue
                    elif number_bad_serp_link == None:
                        number_bad_serp_link += 1
                        print(f'bad serps title: {number_bad_serp_link}')
                        continue
                    
                    if ',' in serp_price: #677,80
                        serp_price = int(serp_price.split(',')[0]) #from "752,95 EUR" to 752
                    elif '.' in serp_price: #1.250,45
                        serp_price = int(serp_price.replace('.', ',').split(',')[0])

                    # print(f'---prod: serp_link <{serp_link}> \n id <{serp_id}> \n serp_title: {serp_title} \n serp_price:{serp_price}')
                    # print(prod_text)
                except:
                    traceback.print_exc()
                    continue

                #if the prod is already in ebay_id = prod exist in prod_db -> ignore it
                if serp_id in ebay_id_list:
                    print(f'this ebay_id already exist {serp_id}')
                    continue
                #if the prod is cheaper than the filter price, ignore that prod
                if serp_price < query_price:
                    print(f'this price <{serp_price}> is too low for the filter price: <{query_price}>')
                    continue
                
                ### begin title filter ###
                #filter by title, split query in words, if all words in serp_title and any excluded kws =  append the serp_link to url_list
                s = query_title.split(' ') #split in words
                n = len(s)
                if n == 1: # when title is only one word
                    if query_title in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r: #if filter returns true == no exc_kws present in serp_title
                            url_list.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}, query_title <{query_title}>')
                        continue
                elif n == 2:
                    #if all the words in query_title are present in serp_title...
                    if s[0] in serp_title and s[1] in serp_title:
                        #append only if there aren't any excluded kw in serp_title
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            url_list.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                    if s[0] in serp_title and s[1] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            url_list.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 3:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            url_list.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 4:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            url_list.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 5:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title and s[4] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        print(r)
                        if r:
                            url_list.append(serp_link)
                            print('appended', serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 6:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title and s[4] in serp_title and s[5] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            url_list.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 7:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title and s[4] in serp_title and s[5] in serp_title and s[6] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            url_list.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 8:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title and s[4] in serp_title and s[5] in serp_title and s[6] in serp_title and s[7] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            url_list.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 9:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title and s[4] in serp_title and s[5] in serp_title and s[6] in serp_title and s[7] in serp_title and s[8] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            url_list.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 10:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title and s[4] in serp_title and s[5] in serp_title and s[6] in serp_title and s[7] in serp_title and s[8] in serp_title and s[9] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            url_list.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                print('END  prod------------------------------------')
                
                # except Exception as e:
                #     print('EXCEPTION..................')
                #     errors_n += 1
                #     print('errors= ', errors_n)
                #     traceback.print_exc()
                #     print(e)
                #     continue
                
            print('return len',len(url_list))
            return url_list
            

        #end title filter 
        #################################################

        start_url =    response.meta["start_url"]
        ebay_id_list = response.meta['ebay_id_list']
        query_category= response.meta['target_category']
        query_title = response.meta['query_title']
        query_price = response.meta['query_price']
        target_category=response.meta['target_category']
        query_attribute_1= response.meta['query_attribute_1']
        query_attribute_2= response.meta['query_attribute_2']
        query_model     = response.meta['query_model']
        query_prod_state = response.meta['query_prod_state']
        available_colors = response.meta['available_colors']

        # excluded_kws =response.meta['excluded_kws']
        # excluded_kws = excluded_kws.split(',')
        
        print('in serp():', '\n',
                        'start_url',start_url,'\n',
                        'ebay_id_list',ebay_id_list,'\n',
                        'query_category',query_category,'\n',
                        'query_title',query_title,'\n',
                        'query_price',query_price,'\n',
                        'available_colors', available_colors, '\n'
                        # 'excluded_kws', excluded_kws,'\n'
                        )

        #get product  webelements from ebay serps
        serp_prods = response.xpath('//li[@class="s-item s-item__pl-on-bottom"]')
        print('-------serp prods count', len(serp_prods))

        #return url list of products that pass title and price
        filtered_prods = filter_price_title(serp_prods, query_title, query_price, ebay_id_list)
        
        #for some reason there're duplicateds, set the list
        print('before', len(filtered_prods))
        filtered_prods = set(filtered_prods)
        print('filtered_prods: ',len(filtered_prods))
        # for u in filtered_prods: print('bbb',u)

        for prod_url in filtered_prods:
            # print('this is the url to pass to parse()',prod_url)
            yield scrapy.Request(url= prod_url, callback=self.parse, meta={
                'start_url':start_url,
                'query':query_title, 
                'target_category':target_category,
                'query_attribute_1':query_attribute_1,
                'query_attribute_2':query_attribute_2,
                'query_model':query_model,
                'query_prod_state':query_prod_state,
                'available_colors':available_colors
                })


    def parse(self, response):        

        def get_category(breadcrumbs):
            from bs4 import BeautifulSoup as bs4
            #breadcrumbs = "<li class=\"vi-VR-brumblnkLst vi-VR-brumb-hasNoPrdlnks\" id=\"vi-VR-brumb-lnkLst\">\n\t\t\t<ul role=\"list\" aria-label=\"En la categoría: \" itemscope=\"\" itemtype=\"https://schema.org/BreadcrumbList\">\n\t\t\t\t\t<li role=\"listitem\" itemprop=\"itemListElement\" itemscope=\"\" itemtype=\"https://schema.org/ListItem\" class=\"bc-w\">\n\t\t\t\t\t\t\t\t\t<a itemprop=\"item\" _sp=\"p2047675.l2706\" href=\"https://www.ebay.es/b/Consolas-y-videojuegos-/1249\" class=\"thrd\">\n\t\t\t\t\t\t\t\t\t\t<span itemprop=\"name\">Consolas y videojuegos</span>\n\t\t\t\t\t\t\t\t\t</a>\n\t\t\t\t\t\t\t\t\t<meta itemprop=\"position\" content=\"1\">\n\t\t\t\t\t\t\t\t</li>\n\t\t\t\t\t\t\t\t<li aria-hidden=\"true\">&gt;</li>\n\t\t\t\t\t\t\t\t<li role=\"listitem\" itemprop=\"itemListElement\" itemscope=\"\" itemtype=\"https://schema.org/ListItem\" class=\"bc-w\">\n\t\t\t\t\t\t\t\t\t<a itemprop=\"item\" _sp=\"p2047675.l2706\" href=\"https://www.ebay.es/b/Videojuegos-/139973\" class=\"scnd\">\n\t\t\t\t\t\t\t\t\t\t<span itemprop=\"name\">Videojuegos</span>\n\t\t\t\t\t\t\t\t\t</a>\n\t\t\t\t\t\t\t\t\t<meta itemprop=\"position\" content=\"2\">\n\t\t\t\t\t\t\t\t</li>\n\t\t\t\t\t\t\t\t<li>&gt;</li>\n\t\t\t\t\t\t<li itemprop=\"itemListElement\" itemscope=\"\" itemtype=\"https://schema.org/ListItem\" class=\"bc-w\">\n\t\t\t\t\t\t\t<a itemprop=\"item\" _sp=\"p2047675.l2644\" href=\"https://www.ebay.es/p/25031842995\" title=\"Ver más Days Gone (Sony PlayStation 4, 2019)\">\n\t\t\t\t\t\t\t\t<span itemprop=\"name\">Ver más Days Gone (Sony PlayStation 4, 2019)</span>\n\t\t\t\t\t\t\t</a>\n\t\t\t\t\t\t\t<meta itemprop=\"position\" content=\"1\">\n\t\t\t\t\t\t</li>\n\t\t\t\t\t</ul>\n\t\t\t</li>"
            try:
                #extract html
                breadcrumbs = response.xpath('//ul[@id="bc"]/li[4]').extract_first()
                #parse html
                soup = bs4(breadcrumbs, 'html.parser')
                #find all the links
                raw = soup.find_all('a')[-2] #-2 because it's allways the penultimate a
                link = raw.get('href')
                #take only text, avoid https:/....
                category_name = link.split('/')[-2] #split using '/' and take only the text of the link
                category_name = category_name.replace('-','')
                return category_name
            except Exception as e:
                print(e)

        # def get_subcategory(response, category):
        #     ''' Assing subcat based on category, if category is phone, seek for model
        #     if category is games seek for platform '''
        #     subcategory='' #define subcategory default value
        #     #if xpath contains platform then the product
        #     if category == "Consolasyvideojuegos":
        #         subcategory = response.xpath('//td[contains(text(),"Plataforma")]/following-sibling::td[1]/span/text()').extract_first()
        #     elif category == "Móviles y Smartphones":
        #         subcategory = response.xpath('//td[contains(text(),"Modèle") or contains(text(),"Modelo") or contains(text(),"Model")]').extract_first()
        #     return subcategory

        # def get_payment_methods(response):
        #     div = response.xpath('id="payDet1"')

        def get_specs(divs):
            ''' if there's a repeated value in the specs it will disorder the output
            FE: if the value 8GB appears 2 times it will only be appended once to list, disordering the other elements'''
            try:
                divs = response.xpath('//div[@class="ux-layout-section__row"]')#.extract()
                des = '' #description or specs
                flag = 0
                for div in divs:
                    labels = div.xpath('.//div[@class="ux-labels-values__labels"]/div/div/span/text()').getall()
                    value_state = div.xpath('.//div[@class="ux-labels-values__values-content"]/div/span/span/span/text()').get()
                    values = div.xpath('.//div[@class="ux-labels-values__values-content"]/div/span/text()').getall()

                    if flag == 0:
                        entry1 = labels[0] +' '+ value_state + '\n'+ '\n'
                        entry2 = labels[1] +' '+ values[0] + '\n'+ '\n'

                        des += entry1
                        des += entry2
                        flag += 1
                    else:
                        try:
                            entry1 = labels[0] +' '+ values[0] + '\n' + '\n'
                            entry2 = labels[1] +' '+ values[1] + '\n' + '\n'
                            des += entry1
                            des += entry2
                        except: pass
                return des
            except Exception as e:
                print(f'error in get_specs() {e}')
                traceback.print_exc()

            #possible fix with a func that checks if there some kw repeated, then use a flag for that kw
            # flag = 0
            # _list = []
            # values_list = []
            # labels_list = []
            # for row in rows:
            #     divs = row.xpath('.//div')
            #     for div in divs:
            #         text = div.xpath('.//span/text()').get()
            #         if text not in _list:
            #             _list.append(text)

            # for text in _list:
            #     # print(text)
            #     if flag %2 == 0:
            #         labels_list.append(text)
            #         # print(flag,'is even', 'label: ',label)
            #     else:
            #         values_list.append(text)
            #         # print(flag,'is odd' ,'value',value)
            #     flag += 1
            # zipped = list(zip(labels_list, values_list))
            # return zipped

        
        #this doesn't work well, in vscode viewer the csv looks fine
        #in excel csv viewer looks with bugs, strange breaklines
        def get_related_links(response):
            related_prod_list = []
            related_links = response.xpath('//li[@class="rtxt"]/parent::ul/li/a/@href').getall()#.extract_first()
            # my_print(len(related_prod_list))
            # for prod in related_links:
            #     related_prod_list.append(prod)
            return related_links

        def get_price(price):
            if price == None:    #notice is different convB..
                price = response.xpath('//span[@id="convbidPrice"]/text()').get()
            if price == None:           #notice the id is conviD and conviN, the're differents
                price = response.xpath('//span[@id="convbinPrice"]/text()').get()
            if price == None:   #this is spanish EUR price if prod is from spain
                price = response.xpath('//span[@class="notranslate"]/text()').extract_first()
            if price == None:
                price = response.xpath('//span[@id="mm-saleDscPrc"]').get()
            if price == None:
                price = response.xpath('//span[@class="notranslate"]').get()
            if price == None:
                price = response.xpath('//span[@class="notranslate "]').get()
            if price == None:
                price = response.xpath('//span[@id="mm-saleDscPrc"]').get()
            return price

        def get_shipping_price():
            try:
                shipping_price = response.xpath('//span[@id="convetedPriceId"]/text()').get()
            except:
                pass

            if shipping_price == None:
                shipping_price = response.xpath('//span[@id="fshippingCost"]/span/text()').get()
            return shipping_price

        def get_ebay_pics(pics_selector):
            #this pics are thumbnails, in filter I change the URL pattern to access the big pics
            pics_list = []
            lis = pics_selector.xpath('.//li')
            for li in lis:
                pic = li.xpath('.//img/@src').get()
                pics_list.append(pic)
            return pics_list

        #################   PARSE FUNCTION   ####################
        # start_url = response.meta["start_url"]
        query = response.meta["query"]
        target_category = response.meta['target_category']
        # prod_specs_html = response.xpath('//div[@class="vim x-about-this-item"]/div//div[@class="ux-layout-section ux-layout-section--features"]/div//div[@class="ux-layout-section__row"]')
        query_attribute_1= response.meta['query_attribute_1']
        query_attribute_2= response.meta['query_attribute_2']
        query_model     = response.meta['query_model']
        query_prod_state = response.meta['query_prod_state']
        available_colors = response.meta['available_colors']
        
        #now target query should be passed in meta from the gaps_file
        # query = get_query_text(start_url)

        #miscelaneus prod data; All Except prod_description
        variable_prod    = response.xpath('//span[@id="sel-msku-variation"]').get()
        title    = response.xpath('//h1[@itemprop="name"]/text()').extract_first()
        returns  = response.xpath('//span[@id="vi-ret-accrd-txt"]/text()').extract_first()
        ebay_article_id  = response.xpath('//div[@id="descItemNumber"]/text()').extract_first()
        prod_url     = response.url
        category     = get_category(response) #taking the last link in breadcrumbs
        ebay_vendor  = response.xpath('//span[@class="mbg-nw"]/text()').extract_first()
        product_state    = response.xpath('//div[@id="vi-itm-cond"]/text()').extract_first()
        #related_links = get_related_links(response) #not using because of a weird bug
        product_sold_out_text = response.xpath('//span[contains(text(),"Este artículo está agotado")]')
        shipping_price= response.xpath('//span[@class="vi-fnf-ship-txt "]/strong/text()').get()
        #if the prod is not from Spain, the xpath for shippment changes
        #shipping_time = response.xpath('//span[@class="vi-acc-del-range"]/b/text()').extract_first()
        shipping_time = response.xpath('//span[@class="vi-del-ship-txt"]/strong[@class="vi-acc-del-range"]/text()').extract_first()
        # when prod is not in spain, this identifies if its shipped to spain or not
        served_area = response.xpath('//span[@itemprop="areaServed"]/text()').get()
        reviews = response.xpath('//div[@class="reviews"]/text()').extract_first()
        seller_votes = response.xpath('//span[@class="mbg-l"]/a/text()').get()
        payment_methods = response.xpath('//div[@id="payDet1"]/div/img/@alt').getall()
        import_taxes = response.xpath('//span[@id="impchCost"]/text()').get()

        subtitle = response.xpath('//div[@id="subTitle"]/span/text()').get()
        if subtitle == None:
            subtitle = response.xpath('//span[@class="topItmCndDscMsg"]/text()').get()
        if subtitle == None:
            subtitle = response.xpath('//div[@id="subTitle"]/text()').get()
        
        # try:
        #     subtitle = subtitle.text
        # except AttributeError: # NoneType has no attr  text
        #     pass

        #translated later in filter.py
        prod_specs_html = response.xpath('//div[@class="ux-layout-section__row"]')
        prod_specs = get_specs(prod_specs_html)
        
        # prod_specs_text = str(prod_specs_html)
        
        #this pic is always available, but not always there are more than one. Use try: to find other pics, if any pics= main
        ebay_main_pic_url = response.xpath('//img[@id="icImg"]/@src').get()
        try: 
            #this pics are thumbnails, later in filter.py I change the url pattern to access the big pictures using the thumbnail URL
            pics_selector = response.xpath('//ul[@class="lst icon"]')[1]
            ebay_pics  = get_ebay_pics(pics_selector)
            ebay_pics.append(ebay_main_pic_url)
        except Exception as e:
            ebay_pics = ebay_main_pic_url
            print(e)
        #this is to replace breaklines that excel don't decode well
        #possible fix: when launching the spider, setting csv output, linebreak = '\n' in arguments
        try:
            reviews = reviews.replace('\n','')
        except:
            pass

        try:
            served_area = served_area.replace('\n','')
        except:
            pass

        #location = response.xpath('//div[@class="iti-eu-bld-gry "]/span/text()').extract()
        article_location = response.xpath('//span[@itemprop="availableAtOrFrom"]/text()').get()

        #depending on article location shows on price or another
        if 'España' not in article_location:
            pass
        if shipping_price == None:
            shipping_price = get_shipping_price()
        if shipping_time == None or '':
                shipping_time=response.xpath('//span[@class="vi-acc-del-range"]/b/text()').extract_first()

        #many prods are from UK, they show price in GBP and EUR,
        # this tries first to take EUR prices with prods with both options
        price = response.xpath('//span[@id="convbinPrice"]/text()').get()
        if price == None or '':
            price = get_price(price) # there're a lot of alternatives, so I made a function

        #this is the iframe with prod_description inside, to get it it's needed another request to get the info inside the iframe
        #scrape all the data and make another request to the iframe's url and pass all the scrapped data as meta.
        #some prods don't have iframe. if iframe == none take the description and yield to finish this item
        iframe_description_url = response.xpath('//div[@id="desc_wrapper_ctr"]/div/iframe/@src').get()
        if iframe_description_url == None:
            iframe_description_url = 'not present'
            # print(f'this prod does not have iframe {prod_url}')
            # yield {'title':title,'price':price, 'query':query,
            #         'shipping_time':shipping_time, 'variable_prod':variable_prod,
            #         'returns':returns,'shipping_price':shipping_price,
            #         'ebay_article_id':ebay_article_id,'prod_url':prod_url,
            #         'ebay_vendor':ebay_vendor,'seller_votes':seller_votes,
            #         'category':category, 'payment_methods':payment_methods,'prod_specs':prod_specs,
            #         'product_state':product_state, #'prod_description':prod_description,
            #         'served_area':served_area,'reviews':reviews,'product_sold_out_text':product_sold_out_text,
            #         'import_taxes':import_taxes, 
            #         'target_category':target_category,
            #         'query_attribute_1':query_attribute_1,
            #         'query_attribute_2':query_attribute_2,
            #         'query_model':query_model,
            #         'query_prod_state':query_prod_state,
            #         'ebay_pics':ebay_pics,
            #         'available_colors':available_colors,
            #         'subtitle':subtitle
            #          }#'related_links':related_links,
        else: #if iframe url exist, make a request to it
            # now call the iframe parser and give it all the info inside meta
            yield Request(url=iframe_description_url, callback=self.iframe,
                        meta={
                            'title':title,
                            'price':price, 
                            'query':query,
                            'shipping_time':shipping_time,
                            'variable_prod':variable_prod,
                            'returns':returns,
                            'shipping_price':shipping_price,
                            'ebay_article_id':ebay_article_id,
                            'prod_url':prod_url,
                            'ebay_vendor':ebay_vendor,
                            'seller_votes':seller_votes,
                            'category':category,
                            'payment_methods':payment_methods,
                            'product_state':product_state,
                            'product_sold_out_text':product_sold_out_text,
                            'served_area':served_area,
                            'reviews':reviews,
                            'prod_specs':prod_specs,
                            'import_taxes':import_taxes,
                            'target_category':target_category,
                            'query_attribute_1':query_attribute_1,
                            'query_attribute_2':query_attribute_2,
                            'query_model':query_model,
                            'query_prod_state':query_prod_state,
                            'ebay_pics':ebay_pics,
                            'available_colors':available_colors,
                            'subtitle':subtitle,
                            'iframe_description_url':iframe_description_url
                            })


    #this scrapes the prod description in iframe
    def iframe(self,response):
        def get_prod_des_text(des_selector):
            prod_description = ''
            paragraphs = des_selector.xpath('.//p/text()').getall()
            for p in paragraphs:
                prod_description += p
                prod_description += ' '
                prod_description += '\n'
            return paragraphs

        
        #this is the prod descrption in the iframe
        # prod_description = response.xpath('//body').extract() #extract to get the html, not the text with get()
        # prod_description = str(prod_description)
        # description_selector = response.xpath('//body/table/text()') #extract to get the html, not the text with get()
        # prod_description = get_prod_des_text(description_selector)
        # prod_description = response.xpath('//body/table/text()').getall() #extract to get the html, not the text with get()
        # prod_description = response.xpath('//div[@id="ds_div"]').extract()
        prod_description = response.xpath('//div[@id="ds_div"]').extract()
        if prod_description == None:
            prod_description = 'not present'

        #get all data from meta
        title   =   response.meta['title']
        price   =   response.meta['price']
        query   =   response.meta['query']
        shipping_time = response.meta['shipping_time']
        variable_prod = response.meta['variable_prod']
        returns =   response.meta['returns']
        shipping_price=response.meta['shipping_price']
        ebay_article_id=response.meta['ebay_article_id']
        prod_url     =  response.meta['prod_url']
        ebay_vendor  =  response.meta['ebay_vendor']
        seller_votes =  response.meta['seller_votes']
        category     =  response.meta['category']
        payment_methods=response.meta['payment_methods']
        product_state  =response.meta['product_state']
        prod_specs   =  response.meta['prod_specs']
        served_area  =  response.meta['served_area']
        reviews  =   response.meta['reviews']
        product_sold_out_text = response.meta['product_sold_out_text']
        import_taxes     =  response.meta['import_taxes']
        target_category  =  response.meta['target_category']
        query_attribute_1=  response.meta['query_attribute_1']
        query_attribute_2=  response.meta['query_attribute_2']
        query_model      =  response.meta['query_model']
        query_prod_state =  response.meta['query_prod_state']
        ebay_pics        =  response.meta['ebay_pics']
        available_colors =  response.meta['available_colors']
        subtitle         =  response.meta['subtitle']
        iframe_description_url = response.meta['iframe_description_url']
        

        yield {'title':title,'price':price, 'query':query,
        'shipping_time':shipping_time,     'variable_prod':variable_prod,
        'returns':returns,                'shipping_price':shipping_price,
        'ebay_article_id':ebay_article_id,'prod_url':prod_url,
        'ebay_vendor':ebay_vendor,      'seller_votes':seller_votes,
        'category':category,            'payment_methods':payment_methods,
        'prod_specs':prod_specs,        'product_state':product_state, 
        'prod_description':prod_description,
        'served_area':served_area,      'reviews':reviews,
        'product_sold_out_text':product_sold_out_text,
        'import_taxes':import_taxes, 
        'target_category':target_category,
        'query_attribute_1':query_attribute_1,
        'query_attribute_2':query_attribute_2,
        'query_model':query_model,
        'query_prod_state':query_prod_state,
        'ebay_pics':ebay_pics,
        'available_colors':available_colors,
        'subtitle':subtitle,
        'iframe_description_url':iframe_description_url
         }#'related_links':related_links,


 # item.add_xpath('warranty', '')
        # item.add_xpath('num_availables','')
        # item.add_xpath('','')

        # item.add_xpath('score', './/div[@class="kVNDLtqL"]/span/text()',
        # MapCompose(self.quitarDolar))
        # # Utilizo Map Compose con funciones anonimas
        # # PARA INVESTIGAR: Que son las funciones anonimas en Python?
        # item.add_xpath('descripcion', '//div[@class="ui_column  "]//div[@class="cPQsENeY"]//text()', # //text() nos permite obtener el texto de todos los hijos
        # MapCompose(lambda i: i.replace('\n', '').replace('\r', '')))


 