''' output file has to be named ebay-spider-output.json '''
import traceback
import scrapy
import csv
from scrapy.http.request import Request
from scrapy.spiders import Spider
from scrapy.selector import Selector
from scrapy.loader.processors import MapCompose
from scrapy.spiders import CrawlSpider, Rule
from scrapy.linkextractors import LinkExtractor
from bs4 import BeautifulSoup as bs4
from scrapy.http import HtmlResponse
import logging
from time import sleep

# logging.basicConfig(level=logging.DEBUG)
# logging.basicConfig(level=logging.INFO)
# logging.basicConfig(level=logging.WARNING)

#for my_print()
from colorama import init
from termcolor import colored
init()

GAPS_FILE = r"C:\Users\HP EliteBook\OneDrive\A_Miscalaneus\Escritorio\Code\git_folder\sm_sys_folder\gaps_file.xlsx"

#created a func, delete when sure
#these are used  in get_queries() and get_query()
# first_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
# last_chunk = '&_sacat=0&rt=nc&LH_BIN=1'


def my_print(text, color='green', mode='normal',**id):
    ''' red,green,yellow,blue // mode="lines" to print a list line by line in BLUE"'''

    if mode == 'normal':
        if color == 'red':
            print(colored(id, 'white','on_red'))
            print(colored(text, 'white','on_red'))
        elif color == 'green':
            print(colored(id, 'white','on_green'),sep="\n")
            print(colored(text, 'white','on_green'),sep="\n")
        elif color == 'yellow':
            print(colored(id, 'blue','on_yellow'))
            print(colored(text, 'blue','on_yellow'))
        elif color == 'blue':
            print(colored(id, 'white','on_blue'))
            print(colored(text, 'white','on_blue'))

    elif mode=='lines':
        for item in text:
            print(colored(item, 'white','on_blue'))

def get_queries():
    """ this takes a file named queries.csv and creates url's to search
    by the scrapper """
    from openpyxl import load_workbook

    data_queries = []

    global first_chunk
    global last_chunk

    #gaps_file
    wb = load_workbook(GAPS_FILE)
    ws = wb.active

    #get the ebay_id_list
    ebay_id_list = ws['F2'].value
    ebay_id_list = ebay_id_list.split(',')
    print(f'---ebay_id_list{ebay_id_list}')

    #get all the rows
    for row in ws.iter_rows(values_only=True, min_row=3):
        query_title = row[0]
        query_quantity = row[1]
        query_category = row[2]
        query_alt_attr = row[3] # FI: "grafito" instead of "negro"
        query_price = row[4] # FI: "grafito" instead of "negro"
        excluded_kws = row[5] # FI: iphone 12 != iphone 12 PRO // exclue those that has MAX or PRO in title

        if query_title == None : continue

        print('gaps_file:',
        query_title,
        query_quantity,
        query_category,
        query_alt_attr)

        #add data to list, return list
        entry = [query_title, query_quantity, query_category, query_alt_attr, query_price, excluded_kws]
        data_queries.append(entry)

    entry = [ebay_id_list, data_queries]
    return entry


def create_url(query_category, query_title):

    #WARNING! use url with 200 results and buy only option

    #use one or other chunks based on target category
    if query_category == 'smartphones':
        # first_url_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_trksid=p2380057.m570.l2632&_nkw='
        # second_url_chunk = '&_sacat=9355&_ipg=200'
        first_url_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_nkw='
        second_url_chunk = '&_sacat=9355&LH_TitleDesc=0&rt=nc&LH_BIN=1&_ipg=200'

    #WARNING! NOT VALID missing  +200 and buy only in url
    elif query_category == 'smartwatches':
        first_url_chunk = 'https://www.ebay.es/sch/i.html?_from=R40&_trksid=p2334524.m570.l2632&_nkw=smartwatch&_sacat=178893&LH_TitleDesc=0&_odkw='
        second_url_chunk = '&_osacat=178893&_ipg=200'

    #create the url
    query_title = query_title.replace(' ', '+')
    query_url = first_url_chunk + query_title + second_url_chunk

    return query_url
#####################              #####################
#####################    CRAWLER   #####################

# class EbaySpiderSpider(CrawlSpider):
class EbaySpiderSpider(scrapy.Spider):
    print('class')
    name = 'ebay-spider'

    def start_requests(self):
        print('inside start_requests')

        # def get_url(self, query_category, query_title):
        #     ''' returns a url based on category and query'''
        #     query_category = 'asdasd'


        # Forma de configurar el USER AGENT en scrapy
        custom_settings = {
            'USER_AGENT': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/71.0.3578.80 Chrome/71.0.3578.80 Safari/537.36'
        }

        # rules = (
        #     Rule(LinkExtractor(restrict_xpaths='//h3[@class="lvtitle"]/a'), callback='parse'),# product links from vendor articles
        #     Rule(LinkExtractor(restrict_xpaths='//div[@class="s-item__image"]/a'), callback='parse'), #product links from ebay search serps results
        #     # Rule(LinkExtractor(restrict_xpaths="//a[@class='pagination__button pagination__next-button']"),callback='parse_tem',),
        # )

        allowed_domains = ['www.ebay.es']

        data_entry = get_queries()
        data_queries = data_entry[1]
        ebay_id_list = data_entry[0]

        for entry in data_queries:
            query_title = entry[0]
            query_quantity = entry[1]
            query_category = entry[2]
            query_alt_attr = entry[3] # FI: "grafito" instead of "negro"
            query_price = entry[4]
            excluded_kws = entry[5]
            excluded_kws = excluded_kws.split(',')

            print('-------------data_query:',
            query_title,
            query_quantity,
            query_category,
            query_alt_attr,
            query_price,
            excluded_kws
            )

            query_url = create_url(query_category, query_title)
            logging.info(f'this is the query url: {query_url}')
            # query_url = 'https://www.ebay.es/sch/15032/i.html?_from=R40&_nkw=iphone+11&LH_TitleDesc=0'

            download_delay = 0.5 # Already set in settings.py

            #for start_url in start_urls:
            yield scrapy.Request(url=query_url, callback=self.serp, meta={'start_url':query_url,
                'query_title':query_title, 'query_quantity':query_quantity, 'query_price':query_price,
                'query_category':query_category, 'query_alt_attr':query_alt_attr,
                'ebay_id_list':ebay_id_list, 'excluded_kws':excluded_kws})

            #test url
            # ['https://www.ebay.es/itm/194207335467?hash=item2d37a8c42b%3Ag%3AguUAAOSw42FgzIP%7E&LH_BIN=1',
            # 'https://www.ebay.es/itm/313519730655?_trkparms=ispr%3D1&hash=item48ff3b6fdf:g:myQAAOSwADxgsNiR&amdata=enc%3AAQAGAAACkPYe5NmHp%252B2JMhMi7yxGiTJkPrKr5t53CooMSQt2orsSB0Xm7k0guzAysaIMfJcbBJmXtVrMUrraCYejHPgiSnRM%252BcPEn4Ba%252FuaUTbl8lT7%252B0FzuQp9fq1%252BaEqI2GrIZ1yCEupo1LkC5Ps%252BjuOFeB2r2Lsx7Tgv6Xg8BGes%252FnYBqR1RvF2IKifEADmNXYARhz%252FPvoIf6Px%252BwdX1qR4FoFyo9pOvOcft26y4wXr3gmafDfdBAIpLYfb3shPeSwpihBZ1aHpZnOVYQs0UcdzNaEr5ymWjQ%252FmSuEuT%252FPI5w4B9TmRM0Ew0XAkT0ZkkCZ88tRSJWxjxHJhnNirYeg5QwOP5SDD2v%252BNlW13CRXzykrdrjpIBSPuDwLFrHCd2bX8QYO9b%252F3DfRqzoIGh8CGedzBX%252BwjbV48W8GEVRPzMlL1AuS4ThusuiQXwKEmNr6IJ2XIX60sxsfPGjGObDeBY88m4z5iWpQ19iSgzU9lJWK0KyneBTKdr8QbUB87P%252FvcrCXYYPtOaXrZxl8Rb%252BVIIl60zujQt0tQO9wZTI8KzUfe108y9PcALaVB7PPmclqJq7S5cNh1XM7TPyVHXAFbTlDs0QyArAa6O8v98Ocagb01cS9PoKC90cI27Omkc0RHb8wN3ufBdGuTj5CnJPRc2dLh3EfCDuJOLrfVYVnPOhDMbWJngGWvjmvsRLGlcEVRkl7U%252BY4wzigH0VsR%252FvxMppiXtm0Sv64NEjLLN1JrH1ldDlcxT2r0%252FFS5ZNWcyEq8vY6WvQmJxwMXS1qAGCZ%252Bj12qEPDSeYqrltk8oVv7wRsLl%252FhdpBEKx3A%252BmeERLrHq9B2hKM0RMi4yfKxdICZQQwIFodjL6GJxFcA3w9luZp8vWPm%7Campid%3APL_CLK%7Cclp%3A2334524',
            # 'https://www.ebay.es/itm/143583232314?_trkparms=ispr%3D1&hash=item216e3a413a:g:WGIAAOSwKU1emitb&amdata=enc%3AAQAGAAACkPYe5NmHp%252B2JMhMi7yxGiTJkPrKr5t53CooMSQt2orsS9d9gBrkNbRQtRn8MVXnYkiJFMm7AxianpWr3Y8yHSWFpKEIfZIOHY7uKPFLdtlfps%252Ffb38sXWfveRXTH%252F7LSmsuktN12xJkLewPieBhWqktIYrSYl%252Ffg40LXhhHKcTHVmDfmoO9AyGJLWb3lTOeP2nMflV%252BA%252FjAGo%252F7ZiAJnyco02PkYfwuc6RDiut4K6lP7x66NmfxVyVqnvm9WxRO3y7JHH1eHAP88w33s6qmV28ey76jcJN1dfRuid24PlfLYVI2xnPbukV0S6V5uPp1KNhdhrW9rkH2pX%252FeV83JY%252B%252FSUtLrdKhCexCn8%252FsG%252FHCF6N6uWZp1kwgTCrcVHMh6GNi63Rp8j58wYxrFIvP8%252BDLKmUg73P7LI1N2529np8yZvbjIRVTL%252BlAYaEHVBwQJgJoZWNj%252Bop1QlgHlr0YaDLTiz491iMeyp8arD3chGIzpdU%252BXd6rSZO%252F3aLKV%252BoFHnrgs2ohVjEeS1EkYQ%252B9SHHcuVie4%252BcLDinCYjWwRMm7vUKblFrT2jvE6%252FIHF1gaN5yqhz3Ju4ycEHTGtgMado0Rhb24rQc%252BHizwzaciSL%252B9%252BdiOFVBFZpxbZgpgPCgrEMR1mnCPaDPZ4mwtsu7okZiVELnwzF1nI6%252B3je9Zqr55Wn6jCV1aaxKTvqIrtXU00%252BII8wnkhckQPoYba7qbkd1e4zZ8KoloHbgZjEVK8kINC7KWVBQH2P2FV0BVouZsjWe8G1nBlY32YRY273Axu1%252FkLoQhm708T5Y9x4HDtt7OkTrF1IsURuR32YREgslaefSM5OyMECfsBmWNPsPZ5LCw0KTWT8%252B5X36XrqbjdCGFgR%7Campid%3APL_CLK%7Cclp%3A2334524,'
            # ]

            # old spider rules
            # Tupla de reglas para direccionar el movimiento de nuestro Crawler a traves de las paginas
            # rules = (
            #     Rule( # Regla de movimiento VERTICAL hacia el detalle de los hoteles
            #         LinkExtractor(
            #             allow=r'/itm/' # Si la URL contiene este patron, haz un requerimiento a esa URL
            #         ), follow=True, callback="parse"), # El callback es el nombre de la funcion que se va a llamar con la respuesta al requerimiento hacia estas URLs
            # )

    #ebay serp with prod links
    def serp(self, response):
        def filter_price_title(serp_prods, query_title, query_price, ebay_id_list, excluded_kws):
            #used in filter to exclude serp_titles with excluded wk
            def excluded_kw_absence(serp_title, excluded_kw):
                '''workaround to exclude kw, iphone 12 pro != iphone 12 // if excluded kw in serp_title: continue'''
                flag = 0

                for kw in excluded_kws:
                    if kw in serp_title:
                        flag +=1 # exc kw found in serp_title

                if flag == 0 :
                    return True # any exc kw in title
                elif flag != 0:
                    return False #some exc kw in title


            #this is the output return
            filtered_urls = []

            for prod in serp_prods:
                # try:
                prod_text = prod.get()
                prod_text = prod_text.encode('utf-8')
                print('serp_prods: prod,', prod_text)
                # serp_link = prod.xpath('//a[@class="s-item__serp_link"]/@href').get()
                # serp_link = prod.xpath('//a[@class="s-item__link"]/@href').get()
                # serp_link = prod.xpath('//a[@target="_blank"]/@href').get()
                # serp_link = prod.xpath('//div/div/div/a[@target="_blank"]/@href').get()
                # serp_link = prod.xpath('//div[@class="s-item__image"]').get()

                # serp_link = prod.xpath('//div[@class="s-item__image"]/a[@tabindex="-1"]/@href').get() #all the same ??
                serp_link = prod.xpath('//a[@class="s-item__link"]/@href').get() 

                print('\n','aaa',serp_link)
                serp_id = serp_link.split('itm/')[1].split('?')[0]
                serp_title = prod.xpath('//h3[@class="s-item__title"]/text()').get()
                serp_title = serp_title.lower()
                serp_price = prod.xpath('//span[@class="s-item__price"]/text()').get()
                serp_price = int(serp_price.split(',')[0]) #from "752,95 EUR" to 752

                # print(f'---prod: link <{serp_link}> id <{serp_id}> \n serp_title: {serp_title} \n serp_price:{serp_price}')

                #if the prod is already in ebay_id = prod exist in prod_db -> ignore it
                if serp_id in ebay_id_list:
                    print(f'this ebay_id already exist {serp_id}')
                    continue
                #if the prod is cheaper than the filter price, ignore that prod
                if serp_price < query_price:
                    print(f'this price <{serp_price}> is too low for the filter price: <{query_price}>')
                    continue

                #filter by title, split query in words, if all words in serp_title and any excluded kws =  append the serp_link to filtered_urls
                s = query_title.split(' ') #split in words
                n = len(s)

                if n == 1: # title is only one word
                    if query_title in serp_title and excluded_kws not in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            filtered_urls.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}, query_title <{query_title}>')
                        continue
                elif n == 2:
                    #if all the words in query_title are present in serp_title...
                    if s[0] in serp_title and s[1] in serp_title:
                        #append only if there aren't any excluded kw in serp_title
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            filtered_urls.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                    if s[0] in serp_title and s[1] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            filtered_urls.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 3:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            filtered_urls.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 4:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            filtered_urls.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 5:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title and s[4] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            filtered_urls.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 6:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title and s[4] in serp_title and s[5] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            filtered_urls.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 7:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title and s[4] in serp_title and s[5] in serp_title and s[6] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            filtered_urls.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 8:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title and s[4] in serp_title and s[5] in serp_title and s[6] in serp_title and s[7] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            filtered_urls.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 9:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title and s[4] in serp_title and s[5] in serp_title and s[6] in serp_title and s[7] in serp_title and s[8] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            filtered_urls.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue
                elif n == 10:
                    if s[0] in serp_title and s[1] in serp_title and s[2] in serp_title and s[3] in serp_title and s[4] in serp_title and s[5] in serp_title and s[6] in serp_title and s[7] in serp_title and s[8] in serp_title and s[9] in serp_title:
                        r = excluded_kw_absence(serp_title, excluded_kws)
                        if r:
                            filtered_urls.append(serp_link)
                    else:
                        print(f'no title match in this prod <{serp_title}>, query_title <{query_title}>')
                        continue

                # except Exception as e:
                #     print('..................')
                #     traceback.print_exc()
                #     print(e)
                #     continue
                
            print('filtered_urls: ',len(filtered_urls))
            # [print('aaaa', url) for url in filtered_urls]

            return filtered_urls

        #end filter def
        #################################################

        start_url = response.meta["start_url"]
        ebay_id_list = response.meta['ebay_id_list']
        query_category = response.meta['query_category']
        query_title = response.meta['query_title']
        query_price = response.meta['query_price']
        excluded_kws = response.meta['excluded_kws']

        print('in serp():', '\n',
                        'start_url',start_url,'\n',
                        'ebay_id_list',ebay_id_list,'\n',
                        'query_category',query_category,'\n',
                        'query_title',query_title,'\n',
                        'query_price',query_price,'\n',
                        'excluded_kws', excluded_kws,'\n')

        #get product  webelements from ebay serps
        serp_prods = response.xpath('//li[@class="s-item s-item__pl-on-bottom"]')
        print('-------serp prods count', len(serp_prods))
        for p in serp_prods: print('---p', p)

        #return url list of products that pass title, price not auction price
        filtered_prods = filter_price_title(serp_prods, query_title, query_price, ebay_id_list, excluded_kws)
        print('filtered_prods: ',len(filtered_prods))

        for prod_url in filtered_prods:
            print('this is the url to pass to parse()',prod_url)
            yield scrapy.Request(url= prod_url, callback=self.parse, meta={'start_url':start_url,'query':query_title })


        #######################################
        # for prod in serp_prods:

        #     # filter_response = filter_price_title(serp_price, serp_title)
        #     filter_url = filter_price_title(prod, query_title, query_price)

        #     #if the prod hasn't the right title or price, continue to next
        #     if filter_url == 'continue':
        #         continue

        #     #else, the prod might be correct, parse it
        #     yield Request(url= filter_url, callback=self.parse, meta={'start_url':start_url})

            ############################  old
            # start_url = response.meta["start_url"]
            # prods_url = response.xpath('//div[@class="s-item__image"]/a[1]/@href').extract()

            # for url in prods_url:
                # yield Request(url=url, callback=self.parse, meta={'start_url':start_url})

    def parse(self, response):
        print('insude of parse:')
        def get_category(breadcrumbs):
            from bs4 import BeautifulSoup as bs4
            #breadcrumbs = "<li class=\"vi-VR-brumblnkLst vi-VR-brumb-hasNoPrdlnks\" id=\"vi-VR-brumb-lnkLst\">\n\t\t\t<ul role=\"list\" aria-label=\"En la categoría: \" itemscope=\"\" itemtype=\"https://schema.org/BreadcrumbList\">\n\t\t\t\t\t<li role=\"listitem\" itemprop=\"itemListElement\" itemscope=\"\" itemtype=\"https://schema.org/ListItem\" class=\"bc-w\">\n\t\t\t\t\t\t\t\t\t<a itemprop=\"item\" _sp=\"p2047675.l2706\" href=\"https://www.ebay.es/b/Consolas-y-videojuegos-/1249\" class=\"thrd\">\n\t\t\t\t\t\t\t\t\t\t<span itemprop=\"name\">Consolas y videojuegos</span>\n\t\t\t\t\t\t\t\t\t</a>\n\t\t\t\t\t\t\t\t\t<meta itemprop=\"position\" content=\"1\">\n\t\t\t\t\t\t\t\t</li>\n\t\t\t\t\t\t\t\t<li aria-hidden=\"true\">&gt;</li>\n\t\t\t\t\t\t\t\t<li role=\"listitem\" itemprop=\"itemListElement\" itemscope=\"\" itemtype=\"https://schema.org/ListItem\" class=\"bc-w\">\n\t\t\t\t\t\t\t\t\t<a itemprop=\"item\" _sp=\"p2047675.l2706\" href=\"https://www.ebay.es/b/Videojuegos-/139973\" class=\"scnd\">\n\t\t\t\t\t\t\t\t\t\t<span itemprop=\"name\">Videojuegos</span>\n\t\t\t\t\t\t\t\t\t</a>\n\t\t\t\t\t\t\t\t\t<meta itemprop=\"position\" content=\"2\">\n\t\t\t\t\t\t\t\t</li>\n\t\t\t\t\t\t\t\t<li>&gt;</li>\n\t\t\t\t\t\t<li itemprop=\"itemListElement\" itemscope=\"\" itemtype=\"https://schema.org/ListItem\" class=\"bc-w\">\n\t\t\t\t\t\t\t<a itemprop=\"item\" _sp=\"p2047675.l2644\" href=\"https://www.ebay.es/p/25031842995\" title=\"Ver más Days Gone (Sony PlayStation 4, 2019)\">\n\t\t\t\t\t\t\t\t<span itemprop=\"name\">Ver más Days Gone (Sony PlayStation 4, 2019)</span>\n\t\t\t\t\t\t\t</a>\n\t\t\t\t\t\t\t<meta itemprop=\"position\" content=\"1\">\n\t\t\t\t\t\t</li>\n\t\t\t\t\t</ul>\n\t\t\t</li>"
            try:
                #extract html
                breadcrumbs = response.xpath('//ul[@id="bc"]/li[4]').extract_first()

                #parse html
                soup = bs4(breadcrumbs, 'html.parser')

                #find all the links
                raw = soup.find_all('a')[-2] #-2 because it's allways the penultimate a
                link = raw.get('href')

                #take only text, avoid https:/....
                category_name = link.split('/')[-2] #split using '/' and take only the text of the link
                category_name = category_name.replace('-','')
                return category_name

            except Exception as e:
                print(e)

        def get_subcategory(response, category):
            ''' Assing subcat based on category, if category is phone, seek for model
            if category is games seek for platform '''

            subcategory='' #define subcategory default value
            #if xpath contains platform then the product
            if category == "Consolasyvideojuegos":
                subcategory = response.xpath('//td[contains(text(),"Plataforma")]/following-sibling::td[1]/span/text()').extract_first()

            elif category == "Móviles y Smartphones":
                subcategory = response.xpath('//td[contains(text(),"Modèle") or contains(text(),"Modelo") or contains(text(),"Model")]').extract_first()

            return subcategory

        def get_payment_methods(response):

            div = response.xpath('id="payDet1"')

        def get_specs(html_data):
            '''Takes html and convert it into a dict of specs key:value,
            like Brand: Samsung'''
            from bs4 import BeautifulSoup

            #clean html
            html_data = html_data.replace('\n','').replace('\t','')

            soup = BeautifulSoup(html_data,'lxml')
            raw_data = soup.find_all('td')
            data_list = []

            for item in raw_data:
                data_list.append(item.get_text())

            #create two lists and an index to identify later even and odds
            odd_keys = []
            even_values = []
            index = 1

            #create 2 lists one with all the keys and the other with all the values
            for item in data_list:
                item = item.strip()# delete white spaces
                if index % 2 == 0:
                    even_values.append(item)
                    index += 1
                else:
                    odd_keys.append(item)
                    index += 1

            #zip and dict the keys
            zip_obj = zip(odd_keys,even_values)
            specs = dict(zip_obj)
            return specs

        #this doesn't work well, in vscode viewer the csv looks fine
        #in excel csv viewer looks with bugs, strange breaklines
        def get_related_links(response):
            related_prod_list = []
            related_links = response.xpath('//li[@class="rtxt"]/parent::ul/li/a/@href').getall()#.extract_first()
            # my_print(len(related_prod_list))
            # for prod in related_links:
            #     related_prod_list.append(prod)
            return related_links

        def get_price(price):
            if price == None:    #notice is different convB..
                price = response.xpath('//span[@id="convbidPrice"]/text()').get()

            if price == None:           #notice the id is conviD and conviN, the're differents
                price = response.xpath('//span[@id="convbinPrice"]/text()').get()

            if price == None:   #this is spanish EUR price if prod is from spain
                price = response.xpath('//span[@class="notranslate"]/text()').extract_first()

            if price == None:
                price = response.xpath('//span[@id="mm-saleDscPrc"]').get()

            if price == None:
                price = response.xpath('//span[@class="notranslate"]').get()

            if price == None:
                price = response.xpath('//span[@class="notranslate "]').get()

            if price == None:
                price = response.xpath('//span[@id="mm-saleDscPrc"]').get()
            return price

        def get_shipping_price():
            #if shipping_price == None:
            try:
                shipping_price = response.xpath('//span[@id="convetedPriceId"]/text()').get()
            except:
                pass

            if shipping_price == None:
                shipping_price = response.xpath('//span[@id="fshippingCost"]/span/text()').get()

            return shipping_price

        #################   PARSE FUNCTION   ####################
        #this is the iframe with prod_description inside, to get it it's needed another request to get the info inside the iframe
        #scrape all the data and make another request to the iframe's url and pass all the scrapped data as meta.
        iframe_description_url = response.xpath('//div[@id="desc_wrapper_ctr"]/div/iframe/@src').get()

        #get query url from meta, use a function to get the query text from the origin url
        start_url = response.meta["start_url"]
        query = response.meta["query"]
        #now target query should be passed in meta from the gaps_file
        # query = get_query_text(start_url)

        #miscelaneus prod data; All Except prod_description
        variable_prod    = response.xpath('//span[@id="sel-msku-variation"]').get()
        title    = response.xpath('//h1[@itemprop="name"]/text()').extract_first()
        returns  = response.xpath('//span[@id="vi-ret-accrd-txt"]/text()').extract_first()
        ebay_article_id  = response.xpath('//div[@id="descItemNumber"]/text()').extract_first()
        prod_url     = response.url
        category     = get_category(response) #taking the last link in breadcrumbs
        ebay_vendor  = response.xpath('//span[@class="mbg-nw"]/text()').extract_first()
        product_state    = response.xpath('//div[@id="vi-itm-cond"]/text()').extract_first()
        #related_links = get_related_links(response) #not using because of a weird bug
        product_sold_out_text = response.xpath('//span[contains(text(),"Este artículo está agotado")]')
        shipping_price= response.xpath('//span[@class="vi-fnf-ship-txt "]/strong/text()').get()
        #if the prod is not from Spain, the xpath for shippment changes
        #shipping_time = response.xpath('//span[@class="vi-acc-del-range"]/b/text()').extract_first()
        shipping_time = response.xpath('//span[@class="vi-del-ship-txt"]/strong[@class="vi-acc-del-range"]/text()').extract_first()
        # when prod is not in spain, this identifies if its shipped to spain or not
        served_area = response.xpath('//span[@itemprop="areaServed"]/text()').get()
        reviews = response.xpath('//div[@class="reviews"]/text()').extract_first()
        seller_votes = response.xpath('//span[@class="mbg-l"]/a/text()').get()
        payment_methods = response.xpath('//div[@id="payDet1"]/div/img/@alt').getall()
        prod_specs_html = response.xpath('//div[@class="itemAttr"]/div/table').get()
        prod_specs_text = str(prod_specs_html)
        prod_specs = get_specs(prod_specs_text)
        import_taxes = response.xpath('//span[@id="impchCost"]/text()').get()
        #this is to replace breaklines that excel don't decode well
        try:
            reviews = reviews.replace('\n','')
        except:
            pass

        try:
            served_area = served_area.replace('\n','')
        except:
            pass

        #location = response.xpath('//div[@class="iti-eu-bld-gry "]/span/text()').extract()
        article_location = response.xpath('//span[@itemprop="availableAtOrFrom"]/text()').get()

        #depending on article location shows on price or another
        if 'España' not in article_location:
            pass

        if shipping_price == None:
            shipping_price = get_shipping_price()

        if shipping_time == None or '':
                shipping_time=response.xpath('//span[@class="vi-acc-del-range"]/b/text()').extract_first()

        #many prods are from UK, they show price in GBP and EUR,
        # this tries first to take EUR prices with prods with both options
        price = response.xpath('//span[@id="convbinPrice"]/text()').get()
        if price == None or '':
            price = get_price(price) # there're a lot of alternatives, so I made a function

        # now call the iframe parser and give it all the info inside meta
        yield Request(url=iframe_description_url, callback=self.iframe,
                    meta={
                        'title':title,'price':price, 'query':query,
                        'shipping_time':shipping_time, 'variable_prod':variable_prod,
                        'returns':returns,'shipping_price':shipping_price,
                        'ebay_article_id':ebay_article_id,'prod_url':prod_url,
                        'ebay_vendor':ebay_vendor,'seller_votes':seller_votes,
                        'category':category, 'payment_methods':payment_methods,
                        'product_state':product_state, 'product_sold_out_text':product_sold_out_text,
                        'served_area':served_area,'reviews':reviews, 'prod_specs':prod_specs,
                        'import_taxes':import_taxes,
                        })


    #this scrapes the prod description in iframe
    def iframe(self,response):


        #this is the prod descrption in the iframe
        prod_description = response.xpath('//body').extract() #extract to get the html, not the text with get()
        prod_description = str(prod_description)

        #get all data from meta
        title = response.meta['title']
        price = response.meta['price']
        query = response.meta['query']
        shipping_time = response.meta['shipping_time']
        variable_prod = response.meta['variable_prod']
        returns = response.meta['returns']
        shipping_price = response.meta['shipping_price']
        ebay_article_id = response.meta['ebay_article_id']
        prod_url = response.meta['prod_url']
        ebay_vendor = response.meta['ebay_vendor']
        seller_votes = response.meta['seller_votes']
        category = response.meta['category']
        payment_methods = response.meta['payment_methods']
        product_state = response.meta['product_state']
        prod_specs = response.meta['prod_specs']
        served_area = response.meta['served_area']
        reviews = response.meta['reviews']
        product_sold_out_text = response.meta['product_sold_out_text']
        import_taxes = response.meta['import_taxes']

        yield {'title':title,'price':price, 'query':query,
        'shipping_time':shipping_time, 'variable_prod':variable_prod,
        'returns':returns,'shipping_price':shipping_price,
        'ebay_article_id':ebay_article_id,'prod_url':prod_url,
        'ebay_vendor':ebay_vendor,'seller_votes':seller_votes,
        'category':category, 'payment_methods':payment_methods,'prod_specs':prod_specs,
        'product_state':product_state, 'prod_description':prod_description,
        'served_area':served_area,'reviews':reviews,'product_sold_out_text':product_sold_out_text,
        'import_taxes':import_taxes,
        #'related_links':related_links,
         }


 # item.add_xpath('warranty', '')
        # item.add_xpath('num_availables','')
        # item.add_xpath('','')

        # item.add_xpath('score', './/div[@class="kVNDLtqL"]/span/text()',
        # MapCompose(self.quitarDolar))
        # # Utilizo Map Compose con funciones anonimas
        # # PARA INVESTIGAR: Que son las funciones anonimas en Python?
        # item.add_xpath('descripcion', '//div[@class="ui_column  "]//div[@class="cPQsENeY"]//text()', # //text() nos permite obtener el texto de todos los hijos
        # MapCompose(lambda i: i.replace('\n', '').replace('\r', '')))


